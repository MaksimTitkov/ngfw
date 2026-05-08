from fastapi import APIRouter, Request, Depends, Query, Form, Body, UploadFile, File, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, delete as sa_delete, cast, Text, or_
from sqlalchemy.orm import selectinload
from app.db.session import get_db, async_session
from app.db.models import CachedRule, Folder, CachedObject, DeviceMeta, NatFolder, CachedNatRule, CachedLog, ChangeLog, RuleTemplate, CachedAnalysis, ScheduledTask, ManagementSystem
from app.services.crypto import encrypt_pw, decrypt_pw
from app.services.deploy_service import DeployService
from app.services.analyzer_service import run_analysis
from app.services.sync_service import SyncService
from app.i18n import base_ctx, get_lang
from app.version import VERSION, RELEASE_CHANNEL
from app.infrastructure.ngfw_client import NGFWClient
from fastapi.templating import Jinja2Templates
from typing import Any, Dict, List, Optional
from pydantic import BaseModel
import asyncio
import logging
import uuid
import math
import json
import csv
import io
from datetime import datetime, timezone, timedelta
import yaml
import openpyxl
from urllib.parse import quote
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO)

LOG_TTL_HOURS = 24  # must match main.py
logger = logging.getLogger(__name__)

# ── SPA helper: serve Vue SPA if built, else None ────────────────────────
import os as _os
_SPA_INDEX = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "..", "static", "dist", "index.html")

def _try_spa():
    return None  # Jinja2 templates are used for all routes

router = APIRouter()
templates = Jinja2Templates(directory="app/templates")

PROTOCOL_MAP = {1: "ICMP", 6: "TCP", 17: "UDP", 47: "GRE", 50: "ESP", 51: "AH"}
GLOBAL_NAME_MAP = {}

# --- HELPER: РџСЂРѕРІРµСЂРєР° Р°РІС‚РѕСЂРёР·Р°С†РёРё ---
def get_current_user(request: Request):
    user = request.session.get("user")
    if not user:
        return None
    return user


async def _log_change(
    db,
    user: dict,
    action: str,
    entity_type: str,
    entity_name: str = "",
    entity_id: str = "",
    device_group_id: str = "",
    detail: str = "",
):
    entry = ChangeLog(
        username        = user.get("username", "unknown"),
        device_group_id = device_group_id or None,
        entity_type     = entity_type,
        entity_id       = entity_id or None,
        entity_name     = entity_name or None,
        action          = action,
        detail          = detail or None,
    )
    db.add(entry)


# --- РњРћР”Р•Р›Р Р”РђРќРќР«РҐ (Pydantic) ---
class AuthRequest(BaseModel):
    host: str
    username: str
    password: str
    device_id: Optional[str] = None

class ReorderRequest(BaseModel):
    folder_id: str
    rule_ids: List[str]

# --- HELPER FUNCTIONS (FORMATTERS) ---
def format_obj_details(obj: CachedObject) -> str:
    d = obj.data or {}
    cat = obj.category
    typ = str(obj.type).lower()
    
    members = d.get('members', [])
    is_group = 'group' in typ or bool(members)
    
    mem_badge = ""
    if members:
        count = len(members)
        member_names = []
        for m_id in members[:15]:
            name = GLOBAL_NAME_MAP.get(m_id, m_id)
            member_names.append(name)
        tooltip_content = chr(10).join(member_names)
        if count > 15: tooltip_content += f"\n...and {count-15} more"
        tooltip = f" title='{tooltip_content.replace(chr(39), '')}'"
        mem_badge = f" <span class='badge bg-primary text-white ms-2' data-bs-toggle='tooltip' data-bs-placement='right' {tooltip}><i class='fa-solid fa-layer-group me-1'></i>{count} items</span>"
    elif is_group:
        raw_keys = list(d.get('_raw_debug', {}).keys())
        mem_badge = f" <span class='text-muted small fst-italic' title='Keys: {raw_keys}'>[Empty Group]</span>"

    if cat == 'net':
        val = d.get('value') or d.get('inet') or d.get('fqdn') or d.get('address')
        if not val:
            start = d.get('start') or d.get('startIp') or d.get('from')
            end = d.get('end') or d.get('endIp') or d.get('to')
            if start and end:
                val = f"{start} - {end}"
        val = val or ""
        if is_group and not val: return mem_badge
        return f"{val}{mem_badge}" if val else (mem_badge or "-")
        
    elif cat == 'service':
        if is_group and not d.get('protocol'): return mem_badge or "-"
        proto_num = d.get('protocol')
        proto_str = PROTOCOL_MAP.get(int(proto_num), str(proto_num)) if str(proto_num).isdigit() else str(proto_num)
        proto_display = f"{proto_str} ({proto_num})" if str(proto_num).isdigit() and proto_str != str(proto_num) else str(proto_num) if proto_num else ""
        
        ports = []
        dst = d.get('dstPorts') or d.get('port')
        if isinstance(dst, str):
            try: dst = json.loads(dst.replace("'", '"'))
            except: ports.append(dst)
        if isinstance(dst, list):
            for p in dst:
                if isinstance(p, dict):
                    if 'singlePort' in p: ports.append(str(p['singlePort'].get('port', '')))
                    elif 'portRange' in p: ports.append(f"{p['portRange'].get('from', '')}-{p['portRange'].get('to', '')}")
                elif isinstance(p, (str, int)): ports.append(str(p))
        elif isinstance(dst, dict):
            if 'singlePort' in dst: ports.append(str(dst['singlePort'].get('port', '')))
            elif 'portRange' in dst: ports.append(f"{dst['portRange'].get('from', '')}-{dst['portRange'].get('to', '')}")
        elif isinstance(dst, (int, str)) and dst: ports.append(str(dst))
            
        ports_display = ", ".join([p for p in ports if p])
        parts = []
        if proto_display: parts.append(f"<b>Proto:</b> <span class='text-secondary'>{proto_display}</span>")
        if ports_display: parts.append(f"<b>Dst:</b> <span class='font-monospace'>{ports_display}</span>")
        res = " | ".join(parts) if parts else "-"
        return res + mem_badge
        
    elif cat == 'urlcat':
        raw = d.get('_raw_debug') or d
        urls = raw.get('urls') or d.get('urls') or []
        if not isinstance(urls, list):
            urls = []
        count = len(urls)
        if count == 0:
            return "<span style='color:#475569;font-style:italic'>No URLs defined</span>"
        preview = ", ".join(str(u) for u in urls[:3])
        if count > 3:
            preview += f" <span style='color:#475569'>+{count-3} more</span>"
        return f"<span style='font-family:monospace;font-size:11px'>{preview}</span>"

    else:
        return str(d.get('value') or d.get('name') or "-")

templates.env.globals['format_obj_details'] = format_obj_details

def rule_to_dict(rule: CachedRule, object_map: Dict[str, CachedObject], device_group_id: str = "") -> Dict[str, Any]:
    d = rule.data or {}

    def _extract_items(field_key: str) -> list:
        """Return list of (id, name_or_none) tuples from a SecurityRule field."""
        section = d.get(field_key)
        if not section:
            return []
        kind = section.get('kind', '')
        if 'ANY' in kind:
            return []
        objects = section.get('objects', [])
        if isinstance(objects, dict):          # OptionalStringArray: {array: [...]}
            objects = objects.get('array', [])
        if not isinstance(objects, list):
            return []
        result = []
        for item in objects:
            if not isinstance(item, dict):
                continue
            # ObjectZone is direct: {id, name, ...}
            if 'id' in item:
                result.append((item['id'], item.get('name')))
                continue
            # NetworkObject / ServiceItem is wrapped: {networkIpAddress: {id, name, ...}}
            for k, v in item.items():
                if isinstance(v, dict) and 'id' in v:
                    result.append((v['id'], v.get('name')))
                    break
        return result

    def resolve_zone(field_key: str) -> str:
        items = _extract_items(field_key)
        if not items:
            return "Any"
        tags = []
        for uid, embedded_name in items[:4]:
            name = embedded_name or (object_map[uid].name if uid in object_map else uid[:8])
            tags.append(
                f"<span class='obj-tag zone' title='{name}'"
                f" data-ext-id='{uid}' data-dg='{device_group_id}'>{name}</span>"
            )
        extra = len(items) - 4
        result = "".join(tags)
        if extra > 0:
            result += f"<span class='obj-tag overflow'>+{extra}</span>"
        return result

    def resolve_objects(field_key: str, css_class: str = '') -> str:
        items = _extract_items(field_key)
        if not items:
            return "Any"
        tags = []
        for uid, embedded_name in items[:4]:
            cached = object_map.get(uid)
            name = (cached.name if cached else embedded_name) or uid[:8]
            cls = f"obj-tag {css_class}".strip()
            tags.append(
                f"<span class='{cls}' title='{name}'"
                f" data-ext-id='{uid}' data-dg='{device_group_id}'>{name}</span>"
            )
        extra = len(items) - 4
        result = "".join(tags)
        if extra > 0:
            result += f"<span class='obj-tag overflow'>+{extra}</span>"
        return result

    def _raw_ids(field_key: str) -> List[str]:
        return [uid for uid, _ in _extract_items(field_key)]

    raw_action = d.get('action', 'allow')
    action = raw_action.split('_')[-1].title() if '_' in raw_action else raw_action

    return {
        "id":            rule.id,
        "ext_id":        rule.ext_id,
        "name":          rule.name,
        "folder_id":     rule.folder_id,
        "description":   d.get('description', ''),
        "log_mode":      d.get('logMode', 'SECURITY_RULE_LOG_MODE_AT_RULE_HIT'),
        # rendered HTML for table display
        "src_zone":      resolve_zone('sourceZone'),
        "src_net":       resolve_objects('sourceAddr'),
        "dst_zone":      resolve_zone('destinationZone'),
        "dst_net":       resolve_objects('destinationAddr'),
        "service":       resolve_objects('service', 'service'),
        "application":   resolve_objects('application', 'app'),
        "url_category":  resolve_objects('urlCategory', 'url'),
        "action":        action,
        "enabled":       d.get('enabled', True),
        "is_modified":   rule.is_modified or False,
        "modified_at":   rule.modified_at or "",
        # raw ID arrays for pre-populating edit modal
        "src_zone_ids":  _raw_ids('sourceZone'),
        "dst_zone_ids":  _raw_ids('destinationZone'),
        "src_net_ids":   _raw_ids('sourceAddr'),
        "dst_net_ids":   _raw_ids('destinationAddr'),
        "service_ids":   _raw_ids('service'),
        "app_ids":       _raw_ids('application'),
        "url_cat_ids":   _raw_ids('urlCategory'),
        "user_ids":      _raw_ids('sourceUser'),
        "ips_profile_id":  ((d.get('ipsProfile')  or {}).get('id') or ''),
        "av_profile_id":   ((d.get('avProfile')   or {}).get('id') or ''),
        "icap_profile_id": ((d.get('icapProfile') or {}).get('id') or ''),
    }

# --- AUTH ROUTES ---

@router.get("/login", response_class=HTMLResponse)
async def login_page(request: Request):
    return templates.TemplateResponse(request, "login.html", base_ctx(request))

def _require_write(request: Request):
    user = request.session.get("user")
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    if user.get("role") == "ro":
        raise HTTPException(status_code=403, detail="Read-only access")


@router.post("/login")
async def login_action(
    request: Request,
    host: str = Form(default=""),
    username: str = Form(...),
    password: str = Form(...),
    ro: str = Form(default=""),
    db: AsyncSession = Depends(get_db),
):
    role = "ro" if ro == "1" else "rw"

    # RO without host: verify credentials against all configured SUs, grant access to all
    if role == "ro" and not host:
        su_all_res = await db.execute(select(ManagementSystem).where(ManagementSystem.is_active == True))  # noqa: E712
        su_all = su_all_res.scalars().all()
        if not su_all:
            return templates.TemplateResponse(request, "login.html", base_ctx(request,
                error="No configured systems found. Add a system using Full Access first."))
        matched_su = None
        for su_candidate in su_all:
            try:
                tmp = NGFWClient(su_candidate.host, verify_ssl=getattr(su_candidate, 'verify_ssl', False))
                await tmp.login(username, password)
                await tmp.close()
                matched_su = su_candidate
                break
            except Exception:
                pass
        if not matched_su:
            return templates.TemplateResponse(request, "login.html", base_ctx(request,
                error="Invalid credentials for all configured systems."))
        request.session["user"] = {
            "host": matched_su.host, "username": username, "password": password,
            "role": "ro",
            # su_id intentionally omitted → pages show all SUs
        }
        logger.info(f"User {username} logged in (role=ro, all SUs).")
        return RedirectResponse(url="/", status_code=303)

    if not host:
        return templates.TemplateResponse(request, "login.html", base_ctx(request, error="Host is required"))

    client = NGFWClient(host, verify_ssl=False)
    try:
        await client.login(username, password)
        await client.close()

        # Find or create ManagementSystem record
        su_res = await db.execute(
            select(ManagementSystem).where(
                ManagementSystem.host == host,
                ManagementSystem.username == username,
            )
        )
        su = su_res.scalar_one_or_none()
        if su is None:
            su = ManagementSystem(
                id=str(uuid.uuid4()),
                name=host,
                host=host,
                username=username,
                password_enc=encrypt_pw(password),
            )
            db.add(su)
        else:
            su.password_enc = encrypt_pw(password)
            su.is_active = True
        # Tag existing DeviceMeta records that have no su_id yet
        from sqlalchemy import update as sa_update
        await db.execute(
            sa_update(DeviceMeta)
            .where(DeviceMeta.su_id == None)  # noqa: E711
            .values(su_id=su.id)
        )
        await db.commit()

        from app.main import refresh_su_list
        await refresh_su_list()

        request.session["user"] = {
            "host": host, "username": username, "password": password,
            "role": role, "su_id": su.id,
        }
        logger.info(f"User {username} logged in (role={role}, su_id={su.id}).")
        return RedirectResponse(url="/", status_code=303)
    except Exception as e:
        logger.error(f"Login failed: {e}")
        return templates.TemplateResponse(request, "login.html", base_ctx(request, error=f"Connection failed: {e}"))

@router.get("/logout")
async def logout(request: Request):
    request.session.clear()
    return RedirectResponse(url="/login", status_code=303)


# --- MANAGEMENT SYSTEMS (Multi-СУ) ---

async def _build_su_tree(db: AsyncSession, tree: dict) -> dict:
    """Enrich flat device tree with ManagementSystem grouping.
    Returns {su_id: {su_name, su_host, devices: {gid: ...}}}
    Devices without su_id go under key '_unknown'.
    """
    meta_res = await db.execute(
        select(DeviceMeta).options(selectinload(DeviceMeta.su))
    )
    meta_map = {m.device_id: m for m in meta_res.scalars().all()}

    su_tree: dict = {}
    for gid, data in tree.items():
        meta = meta_map.get(gid)
        su = meta.su if meta else None
        key = su.id if su else "_unknown"
        if key not in su_tree:
            su_tree[key] = {
                "su_id": key,
                "su_name": su.name if su else "Unknown",
                "su_host": su.host if su else "",
                "devices": {},
            }
        su_tree[key]["devices"][gid] = data
    return su_tree


@router.get("/api/v1/su")
async def list_su(request: Request, db: AsyncSession = Depends(get_db)):
    _require_write(request)
    res = await db.execute(select(ManagementSystem).order_by(ManagementSystem.created_at))
    items = res.scalars().all()
    return JSONResponse([{
        "id": s.id, "name": s.name, "host": s.host, "username": s.username,
        "verify_ssl": s.verify_ssl, "is_active": s.is_active,
        "last_synced_at": s.last_synced_at.isoformat() if s.last_synced_at else None,
    } for s in items])


@router.post("/api/v1/su")
async def create_su(request: Request, db: AsyncSession = Depends(get_db)):
    _require_write(request)
    body = await request.json()
    host = (body.get("host") or "").strip()
    username = (body.get("username") or "").strip()
    password = body.get("password") or ""
    name = (body.get("name") or host).strip()
    verify_ssl = bool(body.get("verify_ssl", False))

    if not host or not username or not password:
        raise HTTPException(status_code=400, detail="host, username and password are required")

    client = NGFWClient(host, verify_ssl=verify_ssl)
    try:
        await client.login(username, password)
        await client.close()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Connection failed: {e}")

    # Check for duplicate
    dup = await db.execute(select(ManagementSystem).where(
        ManagementSystem.host == host, ManagementSystem.username == username))
    existing = dup.scalar_one_or_none()
    if existing:
        existing.name = name
        existing.password_enc = encrypt_pw(password)
        existing.is_active = True
        await db.commit()
        from app.main import refresh_su_list; await refresh_su_list()
        return JSONResponse({"id": existing.id, "status": "updated"})

    su = ManagementSystem(
        id=str(uuid.uuid4()), name=name, host=host, username=username,
        password_enc=encrypt_pw(password), verify_ssl=verify_ssl,
    )
    db.add(su)
    await db.commit()
    from app.main import refresh_su_list; await refresh_su_list()
    return JSONResponse({"id": su.id, "status": "created"}, status_code=201)


@router.put("/api/v1/su/{su_id}")
async def update_su(su_id: str, request: Request, db: AsyncSession = Depends(get_db)):
    _require_write(request)
    res = await db.execute(select(ManagementSystem).where(ManagementSystem.id == su_id))
    su = res.scalar_one_or_none()
    if not su:
        raise HTTPException(status_code=404, detail="Not found")
    body = await request.json()
    if "name" in body:
        su.name = body["name"].strip()
    if "password" in body and body["password"]:
        su.password_enc = encrypt_pw(body["password"])
    if "verify_ssl" in body:
        su.verify_ssl = bool(body["verify_ssl"])
    await db.commit()
    from app.main import refresh_su_list; await refresh_su_list()
    return JSONResponse({"status": "ok"})


@router.delete("/api/v1/su/{su_id}")
async def delete_su(su_id: str, request: Request, db: AsyncSession = Depends(get_db)):
    _require_write(request)
    user = get_current_user(request)
    if user and user.get("su_id") == su_id:
        raise HTTPException(status_code=400, detail="Cannot delete your active connection")
    res = await db.execute(select(ManagementSystem).where(ManagementSystem.id == su_id))
    su = res.scalar_one_or_none()
    if not su:
        raise HTTPException(status_code=404, detail="Not found")
    await db.execute(
        sa_delete(DeviceMeta).where(DeviceMeta.su_id == su_id)
    )
    await db.delete(su)
    await db.commit()
    from app.main import refresh_su_list; await refresh_su_list()
    return JSONResponse({"status": "deleted"})


@router.post("/api/v1/su/{su_id}/sync")
async def sync_su(su_id: str, request: Request, db: AsyncSession = Depends(get_db)):
    _require_write(request)
    res = await db.execute(select(ManagementSystem).where(ManagementSystem.id == su_id))
    su = res.scalar_one_or_none()
    if not su:
        raise HTTPException(status_code=404, detail="Not found")
    password = decrypt_pw(su.password_enc)
    client = NGFWClient(su.host, verify_ssl=su.verify_ssl)
    try:
        await client.login(su.username, password)
        service = SyncService()
        await service.sync_all(db, client, su_id=su_id)
        from datetime import datetime, timezone
        su.last_synced_at = datetime.now(timezone.utc)
        await db.commit()
        return JSONResponse({"status": "ok"})
    except Exception as e:
        logger.error(f"Sync СУ {su_id} failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        await client.close()


# --- ACTIONS (SYNC / COMMIT / REORDER) ---

@router.post("/sync")
async def sync_data(request: Request, device_id: str = Query(None), db: AsyncSession = Depends(get_db)):
    user = get_current_user(request)
    if not user: return JSONResponse({"status": "error", "message": "Unauthorized"}, status_code=401)

    if device_id:
        # Sync one specific device using current session credentials
        client = NGFWClient(user['host'], verify_ssl=False)
        try:
            await client.login(user['username'], user['password'])
            service = SyncService()
            await service.sync_all(db, client, device_group_id=device_id, su_id=user.get('su_id'))
            return JSONResponse({"status": "ok", "device_id": device_id})
        except Exception as e:
            logger.error(f"Sync failed: {e}")
            return JSONResponse({"status": "error", "message": str(e)}, status_code=500)
        finally:
            await client.close()
    else:
        # Sync ALL registered active ManagementSystems using their stored credentials
        from app.services.crypto import decrypt_pw
        all_su_res = await db.execute(
            select(ManagementSystem).where(ManagementSystem.is_active == True)
        )
        all_su = all_su_res.scalars().all()
        if not all_su:
            # Fallback: no registered СУ — sync using session credentials
            client = NGFWClient(user['host'], verify_ssl=False)
            try:
                await client.login(user['username'], user['password'])
                service = SyncService()
                await service.sync_all(db, client, su_id=user.get('su_id'))
                return JSONResponse({"status": "ok"})
            except Exception as e:
                return JSONResponse({"status": "error", "message": str(e)}, status_code=500)
            finally:
                await client.close()

        async def _sync_one(su):
            """Sync one СУ in its own DB session so parallel runs don't share state."""
            password = decrypt_pw(su.password_enc)
            client = NGFWClient(su.host, verify_ssl=su.verify_ssl)
            try:
                await client.login(su.username, password)
                async with async_session() as session:
                    service = SyncService()
                    await service.sync_all(session, client, su_id=su.id)
                return {"su": su.name, "status": "ok"}
            except Exception as e:
                logger.error(f"Sync {su.name} failed: {e}")
                return {"su": su.name, "status": "error", "message": str(e)}
            finally:
                await client.close()

        results = list(await asyncio.gather(*[_sync_one(su) for su in all_su]))
        return JSONResponse({"status": "ok", "results": results})

@router.post("/commit")
async def commit_changes(request: Request, device_id: str = Form(...), db: AsyncSession = Depends(get_db)):
    _require_write(request)
    user = get_current_user(request)
    if not user: return JSONResponse({"status": "error", "message": "Unauthorized"}, status_code=401)

    client = NGFWClient(user['host'], verify_ssl=False)
    try:
        await client.login(user['username'], user['password'])
        deployer = DeployService()
        await deployer.deploy_device_policy(db, client, device_id)

        # Auto-sync the deployed device so local cache reflects actual NGFW state.
        # skip_changelog=True: positions just changed because WE deployed them — not external changes.
        try:
            service = SyncService()
            await service.sync_all(db, client, device_group_id=device_id, skip_changelog=True)
            logger.info(f"Auto-sync after commit completed for device {device_id}")
        except Exception as sync_err:
            logger.warning(f"Auto-sync after commit failed (deploy succeeded): {sync_err}")

        return JSONResponse({"status": "ok", "synced": True})
    except Exception as e:
        logger.error(f"Commit failed: {e}")
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)
    finally:
        await client.close()

# Р­РќР”РџРћРРќРў Р”Р›РЇ РџР•Р Р•РўРђРЎРљРР’РђРќРРЇ (РСЃРїСЂР°РІР»СЏРµС‚ 404 РїСЂРё Drag&Drop)
@router.post("/api/v1/rules/reorder")
async def reorder_rules(request: Request, data: ReorderRequest, db: AsyncSession = Depends(get_db)):
    _require_write(request)
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error", "message": "Unauthorized"}, status_code=401)

    stmt = select(CachedRule).where(CachedRule.id.in_(data.rule_ids))
    rules = (await db.execute(stmt)).scalars().all()
    rule_map = {r.id: r for r in rules}

    for index, r_id in enumerate(data.rule_ids):
        if r_id in rule_map:
            rule = rule_map[r_id]
            rule.folder_id = data.folder_id
            rule.folder_sort_order = index

    # Log a single reorder entry for the folder
    folder = (await db.execute(select(Folder).where(Folder.id == data.folder_id))).scalar_one_or_none()
    folder_name = folder.name if folder else data.folder_id
    dg = folder.device_group_id if folder else ""
    await _log_change(db, user, "reorder", "rule", folder_name, data.folder_id, dg,
                      f"Reordered {len(data.rule_ids)} rules")

    await db.commit()
    return JSONResponse({"status": "ok"})

# --- !!! Р’РћРў Р­РўРћРў Р­РќР”РџРћРРќРў Р‘Р«Р› РџР РћРџРЈР©Р•Рќ !!! ---
@router.get("/api/v1/rules/folders/tree")
async def get_folders_tree(device_group_id: str = Query(None), db: AsyncSession = Depends(get_db)):
    # Р’С‹Р±РёСЂР°РµРј С‚РѕР»СЊРєРѕ РїР°РїРєРё РЅСѓР¶РЅРѕРіРѕ СѓСЃС‚СЂРѕР№СЃС‚РІР°
    stmt = select(Folder).where(Folder.device_group_id == device_group_id).order_by(Folder.section, Folder.sort_order)
    folders = (await db.execute(stmt)).scalars().all()
    
    result = []
    for f in folders:
        result.append({
            "id": f.id,
            "name": f.name,
            "section": f.section,
            "device_group_id": f.device_group_id
        })
    return JSONResponse(result)

class RuleCreateRequest(BaseModel):
    folder_id: str
    name: str
    action: str
    source_ids: List[str] = []
    dest_ids: List[str] = []
    service_ids: List[str] = []
    source_zone_ids: List[str] = []
    dst_zone_ids: List[str] = []
    app_ids: List[str] = []
    url_cat_ids: List[str] = []
    user_ids: List[str] = []
    ips_profile_id: str = ""
    av_profile_id: str = ""
    icap_profile_id: str = ""

class RuleUpdateRequest(BaseModel):
    rule_id: str
    name: str
    action: str
    enabled: bool = True
    log_mode: str = "SECURITY_RULE_LOG_MODE_AT_RULE_HIT"
    description: str = ""
    source_zone_ids: List[str] = []
    dst_zone_ids: List[str] = []
    source_ids: List[str] = []
    dest_ids: List[str] = []
    service_ids: List[str] = []
    app_ids: List[str] = []
    url_cat_ids: List[str] = []
    user_ids: List[str] = []
    ips_profile_id: str = ""
    av_profile_id: str = ""
    icap_profile_id: str = ""

class ToggleRequest(BaseModel):
    rule_id: str
    enabled: bool

class TransferRequest(BaseModel):
    rule_ids: List[str]
    target_gid: str
    target_folder_id: Optional[str] = None
    mode: str = "copy"

class DeleteRequest(BaseModel):
    rule_ids: List[str]

class AcknowledgeRequest(BaseModel):
    rule_ids: List[str]

class BulkToggleRequest(BaseModel):
    rule_ids: List[str]
    enabled: bool

class BulkActionRequest(BaseModel):
    rule_ids: List[str]
    action: str   # allow / deny / drop / reset_both / reset_server / reset_client

class BulkLogRequest(BaseModel):
    rule_ids: List[str]
    log_mode: str  # LOG_MODE_ENABLED / LOG_MODE_DISABLED / LOG_MODE_IF_SESSIONS_EXISTS

@router.post("/api/v1/rules/create")
async def create_rule(request: Request, data: RuleCreateRequest, db: AsyncSession = Depends(get_db)):
    _require_write(request)
    user = get_current_user(request)
    if not user: return JSONResponse({"status": "error", "message": "Unauthorized"}, status_code=401)
    
    client = NGFWClient(user['host'], verify_ssl=False)
    try:
        await client.login(user['username'], user['password'])
        from app.services.rule_creator import rule_creator
        payload = data.dict()
        await rule_creator.create_rule(db, client, payload)
        return JSONResponse({"status": "ok"})
    except Exception as e:
        logger.error(f"Rule creation failed: {e}")
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)
    finally:
        await client.close()

@router.get("/api/v1/object/resolve")
async def api_resolve_object(ext_id: str = Query(...), device_group_id: str = Query(""), db: AsyncSession = Depends(get_db)):
    """Recursively resolve an object from cached_objects, returning a tree up to 5 levels deep."""
    MAX_DEPTH = 5
    seen: set[str] = set()

    def _extract_value(data: dict) -> str | None:
        raw = data.get("_raw_debug") or {}

        # Plain string: IP, CIDR, etc.
        for key in ("value", "inet", "ip", "address"):
            v = data.get(key) or raw.get(key)
            if v and isinstance(v, str):
                return v

        # IP Range
        start = data.get("start") or raw.get("startIp") or raw.get("start")
        end   = data.get("end")   or raw.get("endIp")   or raw.get("end")
        if isinstance(start, str) and isinstance(end, str):
            return f"{start} – {end}"
        if isinstance(start, str):
            return start

        # FQDN
        for key in ("fqdn", "hostFqdn"):
            v = raw.get(key)
            if v and isinstance(v, str):
                return v

        # Service: simple port field
        port  = data.get("port")     or raw.get("port")
        proto = data.get("protocol") or raw.get("protocol") or ""
        if port and isinstance(port, (str, int)):
            p = f"{proto} {port}".strip() if proto else str(port)
            return p

        # Service: dstPorts list of {from/to} dicts  →  "80", "443", "8080-8090"
        dst_ports = data.get("dstPorts") or raw.get("dstPorts") or raw.get("destinationPorts")
        if dst_ports and isinstance(dst_ports, list):
            parts = []
            for p in dst_ports[:4]:
                if isinstance(p, dict):
                    frm = p.get("from") or p.get("start") or p.get("port") or ""
                    to  = p.get("to")   or p.get("end")   or ""
                    if frm and to and str(frm) != str(to):
                        parts.append(f"{frm}-{to}")
                    elif frm:
                        parts.append(str(frm))
                elif isinstance(p, (str, int)):
                    parts.append(str(p))
            if parts:
                suffix = f" +{len(dst_ports)-4}" if len(dst_ports) > 4 else ""
                proto_prefix = f"{proto} " if proto else ""
                return proto_prefix + ", ".join(parts) + suffix
        if isinstance(dst_ports, str):
            return dst_ports

        return None

    # Derive su_id prefix from device_group_id to scope the lookup
    _su_prefix = device_group_id.split(":", 1)[0] if ":" in device_group_id else None

    def _scoped_uid(raw_uid: str) -> str:
        return f"{_su_prefix}:{raw_uid}" if _su_prefix else raw_uid

    async def _resolve(uid: str, depth: int) -> dict:
        if depth > MAX_DEPTH:
            return {"ext_id": uid, "name": "…", "type": "limit", "value": "(too deep)", "members": []}
        if uid in seen:
            return {"ext_id": uid, "name": "…", "type": "cycle", "value": "(circular)", "members": []}
        seen.add(uid)

        # Try scoped lookup first, then fall back to raw (legacy / unscoped data)
        scoped = _scoped_uid(uid)
        stmt = select(CachedObject).where(CachedObject.ext_id.in_([scoped, uid]))
        obj = (await db.execute(stmt)).scalars().first()
        if not obj:
            return {"ext_id": uid, "name": uid[:12], "type": "unknown", "value": None, "members": []}

        data = obj.data or {}
        member_ids: list[str] = data.get("members") or []
        children = []
        for mid in member_ids:
            children.append(await _resolve(str(mid), depth + 1))

        return {
            "ext_id": obj.ext_id,
            "name":   obj.name,
            "type":   obj.type or "",
            "value":  _extract_value(data),
            "members": children,
        }

    node = await _resolve(ext_id, 0)
    return JSONResponse(node)


@router.get("/api/v1/objects/list")
async def api_get_objects_list(device_group_id: str, category: str = "network", db: AsyncSession = Depends(get_db)):
    # Derive scoped "global" from the device_group_id prefix (handles multi-su mode)
    scoped_global = "global"
    if ":" in device_group_id:
        su_prefix = device_group_id.split(":", 1)[0]
        scoped_global = f"{su_prefix}:global"
    stmt = select(CachedObject.ext_id, CachedObject.name, CachedObject.device_group_id).where(
        CachedObject.device_group_id.in_([device_group_id, scoped_global, "global"]),
        func.lower(CachedObject.category).ilike(category.lower() + "%"),
    ).order_by(CachedObject.name)
    rows = (await db.execute(stmt)).all()
    global_ids = {scoped_global, "global"}
    return JSONResponse([{
        "id": r.ext_id,
        "name": f"[Global] {r.name}" if r.device_group_id in global_ids else r.name,
    } for r in rows])

@router.post("/api/v1/rules/acknowledge")
async def acknowledge_rules(request: Request, data: AcknowledgeRequest, db: AsyncSession = Depends(get_db)):
    """Clear the is_modified flag for the given rule IDs."""
    _require_write(request)
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error", "message": "Unauthorized"}, status_code=401)
    rule_ids = data.rule_ids
    for rid in rule_ids:
        rule = await db.get(CachedRule, rid)
        if rule:
            rule.is_modified = False
            rule.modified_at = None
    await db.commit()
    return JSONResponse({"status": "ok", "cleared": len(rule_ids)})


@router.post("/api/v1/rules/update")
async def update_rule_endpoint(request: Request, data: RuleUpdateRequest, db: AsyncSession = Depends(get_db)):
    _require_write(request)
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error", "message": "Unauthorized"}, status_code=401)

    rule = await db.get(CachedRule, data.rule_id)
    if not rule:
        return JSONResponse({"status": "error", "message": "Rule not found"}, status_code=404)

    folder = await db.get(Folder, rule.folder_id)
    device_group_id = folder.device_group_id if folder else None

    def build_field(ids, kind_any="RULE_KIND_ANY", kind_list="RULE_KIND_LIST"):
        if not ids:
            return {"kind": kind_any, "objects": {"array": []}}
        return {"kind": kind_list, "objects": {"array": list(ids)}}

    def build_user_field(ids):
        if not ids:
            return {"kind": "RULE_USER_KIND_ANY", "objects": {"array": []}}
        return {"kind": "RULE_USER_KIND_LIST", "objects": {"array": list(ids)}}

    action_map = {
        "allow": "SECURITY_RULE_ACTION_ALLOW",
        "drop": "SECURITY_RULE_ACTION_DROP",
        "deny": "SECURITY_RULE_ACTION_DENY",
        "reset_server": "SECURITY_RULE_ACTION_RESET_SERVER",
        "reset_client": "SECURITY_RULE_ACTION_RESET_CLIENT",
        "reset_both": "SECURITY_RULE_ACTION_RESET_BOTH",
    }

    from app.services.sync_service import _unscope
    raw_rule_ext_id = _unscope(rule.ext_id)
    raw_device_group_id = _unscope(device_group_id) if device_group_id else None

    api_payload = {
        "id": raw_rule_ext_id,
        "name": data.name,
        "description": data.description,
        "action": action_map.get(data.action.lower(), "SECURITY_RULE_ACTION_ALLOW"),
        "enabled": data.enabled,
        "logMode": data.log_mode,
        "sourceZone":      build_field(data.source_zone_ids),
        "destinationZone": build_field(data.dst_zone_ids),
        "sourceAddr":      build_field(data.source_ids),
        "destinationAddr": build_field(data.dest_ids),
        "service":         build_field(data.service_ids),
        "application":     build_field(data.app_ids),
        "urlCategory":     build_field(data.url_cat_ids),
        "sourceUser":      build_user_field(data.user_ids),
    }
    if data.ips_profile_id:
        api_payload["ipsProfileId"] = data.ips_profile_id
    if data.av_profile_id:
        api_payload["avProfileId"] = data.av_profile_id
    if data.icap_profile_id:
        api_payload["icapProfileId"] = data.icap_profile_id

    client = NGFWClient(user['host'], verify_ssl=False)
    try:
        await client.login(user['username'], user['password'])
        await client.update_rule(raw_rule_ext_id, api_payload)

        full_data = await client.fetch_single_rule(raw_rule_ext_id, raw_device_group_id)
        if full_data:
            rule.name = full_data.get("name", rule.name)
            rule.data = full_data
            rule.is_modified = False
            rule.modified_at = None

        await _log_change(db, user, "update", "rule", rule.name, rule.ext_id,
                          device_group_id or "", f"Action: {data.action}")
        await db.commit()
        return JSONResponse({"status": "ok"})
    except Exception as e:
        logger.error(f"Rule update failed: {e}")
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)
    finally:
        await client.close()


@router.post("/api/v1/rules/toggle")
async def toggle_rule(request: Request, data: ToggleRequest, db: AsyncSession = Depends(get_db)):
    _require_write(request)
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error", "message": "Unauthorized"}, status_code=401)

    rule = await db.get(CachedRule, data.rule_id)
    if not rule:
        return JSONResponse({"status": "error", "message": "Rule not found"}, status_code=404)

    from app.services.sync_service import _unscope
    raw_ext_id = _unscope(rule.ext_id)
    client = NGFWClient(user['host'], verify_ssl=False)
    try:
        await client.login(user['username'], user['password'])
        await client.update_rule(raw_ext_id, {"id": raw_ext_id, "enabled": data.enabled})
        if rule.data:
            rule.data = {**rule.data, "enabled": data.enabled}
        state = "enabled" if data.enabled else "disabled"
        folder = await db.get(Folder, rule.folder_id)
        dg = folder.device_group_id if folder else ""
        await _log_change(db, user, "toggle", "rule", rule.name, rule.ext_id, dg,
                          f"Set {state}")
        await db.commit()
        return JSONResponse({"status": "ok"})
    except Exception as e:
        logger.error(f"Toggle rule failed: {e}")
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)
    finally:
        await client.close()


@router.post("/api/v1/rules/bulk_toggle")
async def bulk_toggle_rules(request: Request, data: BulkToggleRequest, db: AsyncSession = Depends(get_db)):
    _require_write(request)
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error", "message": "Unauthorized"}, status_code=401)

    stmt = select(CachedRule).where(CachedRule.id.in_(data.rule_ids))
    rules = (await db.execute(stmt)).scalars().all()
    if not rules:
        return JSONResponse({"status": "error", "message": "No rules found"}, status_code=404)

    from app.services.sync_service import _unscope
    client = NGFWClient(user['host'], verify_ssl=False)
    ok_count = 0
    errors = []
    state = "enabled" if data.enabled else "disabled"
    try:
        await client.login(user['username'], user['password'])
        for rule in rules:
            try:
                raw_ext = _unscope(rule.ext_id)
                await client.update_rule(raw_ext, {"id": raw_ext, "enabled": data.enabled})
                if rule.data:
                    rule.data = {**rule.data, "enabled": data.enabled}
                folder = await db.get(Folder, rule.folder_id)
                dg = folder.device_group_id if folder else ""
                await _log_change(db, user, "toggle", "rule", rule.name, rule.ext_id, dg,
                                  f"Bulk: set {state}")
                ok_count += 1
            except Exception as e:
                errors.append(f"{rule.name}: {e}")
        await db.commit()
        return JSONResponse({"status": "ok", "updated": ok_count, "errors": errors})
    except Exception as e:
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)
    finally:
        await client.close()


_ACTION_MAP = {
    "allow":        "SECURITY_RULE_ACTION_ALLOW",
    "deny":         "SECURITY_RULE_ACTION_DENY",
    "drop":         "SECURITY_RULE_ACTION_DROP",
    "reset_both":   "SECURITY_RULE_ACTION_RESET_BOTH",
    "reset_server": "SECURITY_RULE_ACTION_RESET_SERVER",
    "reset_client": "SECURITY_RULE_ACTION_RESET_CLIENT",
}

@router.post("/api/v1/rules/bulk_action")
async def bulk_change_action(request: Request, data: BulkActionRequest, db: AsyncSession = Depends(get_db)):
    _require_write(request)
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error", "message": "Unauthorized"}, status_code=401)

    api_action = _ACTION_MAP.get(data.action.lower())
    if not api_action:
        return JSONResponse({"status": "error", "message": f"Unknown action: {data.action}"}, status_code=400)

    stmt = select(CachedRule).where(CachedRule.id.in_(data.rule_ids))
    rules = (await db.execute(stmt)).scalars().all()

    client = NGFWClient(user['host'], verify_ssl=False)
    ok_count = 0
    errors = []
    from app.services.sync_service import _unscope
    try:
        await client.login(user['username'], user['password'])
        for rule in rules:
            try:
                raw_ext = _unscope(rule.ext_id)
                await client.update_rule(raw_ext, {"id": raw_ext, "action": api_action})
                if rule.data:
                    rule.data = {**rule.data, "action": api_action}
                folder = await db.get(Folder, rule.folder_id)
                dg = folder.device_group_id if folder else ""
                await _log_change(db, user, "update", "rule", rule.name, rule.ext_id, dg,
                                  f"Bulk action → {data.action}")
                ok_count += 1
            except Exception as e:
                errors.append(f"{rule.name}: {e}")
        await db.commit()
        return JSONResponse({"status": "ok", "updated": ok_count, "errors": errors})
    except Exception as e:
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)
    finally:
        await client.close()


@router.post("/api/v1/rules/bulk_log")
async def bulk_change_log(request: Request, data: BulkLogRequest, db: AsyncSession = Depends(get_db)):
    _require_write(request)
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error", "message": "Unauthorized"}, status_code=401)

    stmt = select(CachedRule).where(CachedRule.id.in_(data.rule_ids))
    rules = (await db.execute(stmt)).scalars().all()

    client = NGFWClient(user['host'], verify_ssl=False)
    ok_count = 0
    errors = []
    try:
        await client.login(user['username'], user['password'])
        for rule in rules:
            try:
                await client.update_rule(rule.ext_id, {"id": rule.ext_id, "logMode": data.log_mode})
                if rule.data:
                    rule.data = {**rule.data, "logMode": data.log_mode}
                folder = await db.get(Folder, rule.folder_id)
                dg = folder.device_group_id if folder else ""
                await _log_change(db, user, "update", "rule", rule.name, rule.ext_id, dg,
                                  f"Bulk log → {data.log_mode}")
                ok_count += 1
            except Exception as e:
                errors.append(f"{rule.name}: {e}")
        await db.commit()
        return JSONResponse({"status": "ok", "updated": ok_count, "errors": errors})
    except Exception as e:
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)
    finally:
        await client.close()


@router.get("/api/v1/profiles/list")
async def get_profiles_list(request: Request, profile_type: str, device_group_id: str = ""):
    user = get_current_user(request)
    if not user:
        return JSONResponse([], status_code=401)

    client = NGFWClient(user['host'], verify_ssl=False)
    try:
        await client.login(user['username'], user['password'])
        dgid = device_group_id or None
        if profile_type == 'ips':
            items = await client.get_ips_profiles(dgid)
        elif profile_type == 'av':
            items = await client.get_av_profiles(dgid)
        elif profile_type == 'icap':
            items = await client.get_icap_profiles(dgid)
        else:
            items = []
        return JSONResponse([{"id": p.get("id"), "name": p.get("name", "")} for p in items if p.get("id")])
    except Exception as e:
        logger.error(f"Profile list failed: {e}")
        return JSONResponse([])
    finally:
        await client.close()


@router.post("/api/v1/rules/transfer")
async def transfer_rules(request: Request, data: TransferRequest, db: AsyncSession = Depends(get_db)):
    _require_write(request)
    user = get_current_user(request)
    if not user: return JSONResponse({"status": "error", "message": "Unauthorized"}, status_code=401)

    from app.services.transfer_service import TransferService
    from app.services.sync_service import _unscope

    source_su_id = user.get('su_id')

    # Derive target su_id from the scoped device_group_id prefix
    target_su_id = data.target_gid.split(":", 1)[0] if ":" in data.target_gid else source_su_id

    # Source client — used for delete (move) on the source NGFW
    source_client = NGFWClient(user['host'], verify_ssl=False)
    target_client = source_client
    target_client_owned = False  # track whether we created a separate client

    try:
        await source_client.login(user['username'], user['password'])

        # If target is on a different СУ, open a separate connection
        if target_su_id and target_su_id != source_su_id:
            su_res = await db.execute(select(ManagementSystem).where(ManagementSystem.id == target_su_id))
            target_su = su_res.scalars().first()
            if target_su:
                target_client = NGFWClient(target_su.host, verify_ssl=target_su.verify_ssl)
                await target_client.login(target_su.username, decrypt_pw(target_su.password_enc))
                target_client_owned = True

        t_service = TransferService(db, target_client,
                                    source_su_id=source_su_id,
                                    target_su_id=target_su_id)

        for rid in data.rule_ids:
            try:
                await t_service.transfer_rule(rid, data.target_gid, data.target_folder_id)
                if data.mode == "move":
                    stmt = select(CachedRule).where(CachedRule.id == rid)
                    old_r = (await db.execute(stmt)).scalar_one_or_none()
                    if old_r:
                        await source_client.delete_rule(_unscope(old_r.ext_id))
                        await db.delete(old_r)
            except Exception as re_err:
                logger.error(f"Error transferring rule {rid}: {re_err}")

        await db.commit()
        return JSONResponse({"status": "ok", "conflicts": t_service.newly_created_objects})
    except Exception as e:
        logger.error(f"Transfer failed: {e}")
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)
    finally:
        await source_client.close()
        if target_client_owned:
            await target_client.close()

@router.post("/api/v1/rules/delete")
async def delete_rules(request: Request, data: DeleteRequest, db: AsyncSession = Depends(get_db)):
    _require_write(request)
    user = get_current_user(request)
    client = NGFWClient(user['host'], verify_ssl=False)
    await client.login(user['username'], user['password'])
    stmt = select(CachedRule).where(CachedRule.id.in_(data.rule_ids))
    rules = (await db.execute(stmt)).scalars().all()
    for rule in rules:
        await client.delete_rule(rule.ext_id)
        folder = await db.get(Folder, rule.folder_id)
        dg = folder.device_group_id if folder else ""
        await _log_change(db, user, "delete", "rule", rule.name, rule.ext_id, dg)
        await db.delete(rule)
    await client.close()
    await db.commit()
    return JSONResponse({"status": "ok"})

# --- OTHER ROUTES ---

@router.get("/find-rule", response_class=HTMLResponse)
async def find_rule(
    request: Request,
    rule_name: str = Query(""),
    rule_id: str = Query(""),
    device_group_id: str = Query(""),
    db: AsyncSession = Depends(get_db),
):
    """Redirect from changelog/logs to the correct folder for a rule.
    Search priority: ext_id > name+device > name only.
    """
    if not get_current_user(request):
        return RedirectResponse("/login")
    if not rule_name and not rule_id:
        return RedirectResponse("/")

    rule = None

    # 1. By ext_id — unique per NGFW, no ambiguity
    if rule_id:
        stmt = select(CachedRule).where(CachedRule.ext_id == rule_id)
        rule = (await db.execute(stmt)).scalar_one_or_none()

    # 2. By name + device_group_id — resolves same-name rules on different devices
    if not rule and rule_name and device_group_id:
        stmt = (
            select(CachedRule)
            .join(Folder, CachedRule.folder_id == Folder.id)
            .where(CachedRule.name == rule_name, Folder.device_group_id == device_group_id)
            .limit(1)
        )
        rule = (await db.execute(stmt)).scalar_one_or_none()

    # 3. Fallback: name only
    if not rule and rule_name:
        stmt = select(CachedRule).where(CachedRule.name == rule_name).limit(1)
        rule = (await db.execute(stmt)).scalar_one_or_none()

    display_name = rule_name or (rule.name if rule else "")
    if rule and rule.folder_id:
        return RedirectResponse(f"/?folder_id={rule.folder_id}&highlight_rule={quote(display_name)}")
    return RedirectResponse(f"/?rule_name={quote(display_name)}")


@router.get("/find-nat-rule", response_class=HTMLResponse)
async def find_nat_rule(
    request: Request,
    rule_id: str = Query(""),
    rule_name: str = Query(""),
    device_group_id: str = Query(""),
    db: AsyncSession = Depends(get_db),
):
    """Redirect from changelog to the correct NAT folder."""
    if not get_current_user(request):
        return RedirectResponse("/login")

    rule = None
    if rule_id:
        stmt = select(CachedNatRule).where(CachedNatRule.ext_id == rule_id)
        rule = (await db.execute(stmt)).scalar_one_or_none()
    if not rule and rule_name and device_group_id:
        stmt = (
            select(CachedNatRule)
            .where(CachedNatRule.name == rule_name, CachedNatRule.device_group_id == device_group_id)
            .limit(1)
        )
        rule = (await db.execute(stmt)).scalar_one_or_none()
    if not rule and rule_name:
        stmt = select(CachedNatRule).where(CachedNatRule.name == rule_name).limit(1)
        rule = (await db.execute(stmt)).scalar_one_or_none()

    if rule and rule.folder_id:
        return RedirectResponse(f"/nat?folder_id={rule.folder_id}")
    return RedirectResponse("/nat")


@router.post("/create_folder")
async def create_folder(request: Request, folder_name: str = Form(...), device_id: str = Form(...), section: str = Form(...), db: AsyncSession = Depends(get_db)):
    user = get_current_user(request)
    if not user: return RedirectResponse("/login", status_code=303)
    stmt = select(func.max(Folder.sort_order)).where(Folder.device_group_id == device_id)
    max_sort = (await db.execute(stmt)).scalar() or 0
    new_folder = Folder(
        id=str(uuid.uuid4()), name=folder_name, device_group_id=device_id,
        section=section, sort_order=max_sort + 1, su_id=user.get('su_id'),
    )
    db.add(new_folder)
    await db.commit()
    return RedirectResponse(url=f"/?folder_id={new_folder.id}", status_code=303)

@router.get("/", response_class=HTMLResponse)
async def index(request: Request, folder_id: str = Query(None),
                page: int = Query(1), page_size: int = Query(50),
                hl: str = Query(None),
                hl_shadow: str = Query(None),
                hl_shadowed: str = Query(None),
                hl_primary: str = Query("shadow"),
                db: AsyncSession = Depends(get_db)):
    user = get_current_user(request)
    if not user: return RedirectResponse(url="/login")
    spa = _try_spa()
    if spa: return spa

    # Show ALL devices from ALL СУ — IDs are already scoped ({su_id}:uuid), no collision risk
    folder_stmt = select(Folder).options(selectinload(Folder.rules)).order_by(Folder.device_group_id, Folder.sort_order)
    all_folders = (await db.execute(folder_stmt)).scalars().all()

    device_names = {o.device_id: o.name for o in (await db.execute(select(DeviceMeta))).scalars().all()}

    tree = {}
    first_folder_id = None
    first_device_id = None

    for f in all_folders:
        if not first_folder_id: first_folder_id = str(f.id)
        if not first_device_id: first_device_id = f.device_group_id

        gid = f.device_group_id or "unknown"
        dev_name = device_names.get(gid, f"Device {gid[:8]}")
        if gid not in tree: tree[gid] = {"name": dev_name, "id": gid, "sections": {"pre": [], "post": [], "default": []}}
        f.device_name = dev_name; f.rules_count = len(f.rules)
        sec = f.section.lower() if f.section and f.section.lower() in ['pre', 'post', 'default'] else 'pre'
        tree[gid]["sections"][sec].append(f)

    target_folder = next((f for f in all_folders if str(f.id) == folder_id), None)
    if folder_id and not target_folder:
        folder_id = first_folder_id
        target_folder = next((f for f in all_folders if str(f.id) == folder_id), None)

    selected_fid = folder_id or first_folder_id
    current_device_id = target_folder.device_group_id if target_folder else first_device_id

    dashboard_data = []

    obj_stmt = select(CachedObject.ext_id, CachedObject.name, CachedObject.data)
    class MockObj:
        def __init__(self, ext_id, name, data): self.ext_id = ext_id; self.name = name; self.data = data

    GLOBAL_NAME_MAP.clear()
    object_map = {}
    for row in (await db.execute(obj_stmt)).fetchall():
        obj = MockObj(row.ext_id, row.name, row.data)
        object_map[row.ext_id] = obj
        GLOBAL_NAME_MAP[row.ext_id] = row.name
        # Also index by unscoped UUID so rule data lookups (raw NGFW UUIDs) resolve correctly
        if ":" in row.ext_id:
            raw_id = row.ext_id.split(":", 1)[1]
            object_map.setdefault(raw_id, obj)
            GLOBAL_NAME_MAP.setdefault(raw_id, row.name)

    page_size   = max(10, min(page_size, 200))
    total_rules = 0
    total_pages = 1

    # Determine which ext_id we're jumping to (for auto-page calculation)
    jump_ext_id = hl or (hl_shadowed if hl_primary == "shadowed" else hl_shadow)

    if target_folder:
        dg_id_for_tooltip = target_folder.device_group_id or ""
        rules_sorted = sorted(target_folder.rules, key=lambda x: x.folder_sort_order)
        total_rules  = len(rules_sorted)
        total_pages  = max(1, math.ceil(total_rules / page_size))

        # Auto-jump to the page containing the highlighted rule
        if jump_ext_id:
            for idx, r in enumerate(rules_sorted):
                if r.ext_id == jump_ext_id:
                    page = idx // page_size + 1
                    break

        page        = max(1, min(page, total_pages))
        rules_slice = rules_sorted[(page - 1) * page_size : page * page_size]
        rules_processed = [rule_to_dict(r, object_map, dg_id_for_tooltip) for r in rules_slice]
        dashboard_data.append({"folder": target_folder, "rules": rules_processed})

    filtered_tree = {k: v for k, v in tree.items() if not k.endswith(":global") and k != "global"}
    su_tree = await _build_su_tree(db, filtered_tree)

    return templates.TemplateResponse(request, "index.html", base_ctx(request,
        tree=filtered_tree, su_tree=su_tree, dashboard_data=dashboard_data,
        selected_folder_id=selected_fid, current_device_id=current_device_id,
        user=user,
        pg_page=page, pg_size=page_size,
        pg_total=total_rules, pg_pages=total_pages,
        pg_folder=selected_fid,
        hl=hl, hl_shadow=hl_shadow, hl_shadowed=hl_shadowed,
    ))

@router.get("/objects", response_class=HTMLResponse)
async def list_objects(request: Request, su_id: str = Query(None), device_id: str = Query(None), page: int = Query(1), type_filter: str = Query('net'), db: AsyncSession = Depends(get_db)):
    user = get_current_user(request)
    if not user: return RedirectResponse(url="/login")

    PAGE_SIZE = 150

    # Load all СУ for the top-level selector
    su_list = (await db.execute(select(ManagementSystem).order_by(ManagementSystem.created_at))).scalars().all()

    # Derive su_id from scoped device_id prefix when not explicitly provided
    if not su_id and device_id and ":" in device_id:
        su_id = device_id.split(":")[0]

    # Determine active СУ
    if su_id and any(s.id == su_id for s in su_list):
        selected_su_id = su_id
    else:
        selected_su_id = su_list[0].id if su_list else None

    # Load devices for the selected СУ only
    all_devices = (await db.execute(select(DeviceMeta).order_by(DeviceMeta.name))).scalars().all()
    devices = [d for d in all_devices
               if d.su_id == selected_su_id
               and not d.device_id.endswith(":global")
               and d.device_id != "global"]

    # Determine selected device within the active СУ
    valid_ids = {d.device_id for d in devices}
    if device_id and device_id in valid_ids:
        selected_device_id = device_id
    else:
        selected_device_id = devices[0].device_id if devices else None

    GLOBAL_NAME_MAP.clear()
    for row in (await db.execute(select(CachedObject.ext_id, CachedObject.name))):
        GLOBAL_NAME_MAP[row.ext_id] = row.name
        if ":" in row.ext_id:
            GLOBAL_NAME_MAP.setdefault(row.ext_id.split(":", 1)[1], row.name)

    query = select(CachedObject).where(CachedObject.device_group_id == selected_device_id)
    cat_map = {
        'net':     ['net', 'Network', 'Network Group', 'Host/Network'],
        'service': ['service', 'Service', 'Service Group'],
        'app':     ['app', 'Application'],
        'urlcat':  ['urlcat', 'URL Category'],
        'user':    ['user', 'User', 'User Group'],
        'zone':    ['zone', 'Zone', 'Security Zone'],
    }
    target_cats = cat_map.get(type_filter, [])
    if target_cats: query = query.where(CachedObject.category.in_(target_cats))

    count_stmt = select(func.count()).select_from(query.subquery('t'))
    total_items = (await db.execute(count_stmt)).scalar_one()
    total_pages = max(1, math.ceil(total_items / PAGE_SIZE))

    query = query.order_by(CachedObject.name).offset((page - 1) * PAGE_SIZE).limit(PAGE_SIZE)
    objects = (await db.execute(query)).scalars().all()

    return templates.TemplateResponse(request, "objects.html", base_ctx(request,
        user=user, su_list=su_list, selected_su_id=selected_su_id,
        devices=devices, selected_device_id=selected_device_id,
        objects=objects, type_filter=type_filter, page=page,
        total_pages=total_pages, total_items=total_items,
    ))


# ===========================================================================
#  OBJECTS CRUD
# ===========================================================================

class ObjectCreateRequest(BaseModel):
    device_group_id: str
    obj_type: str        # net_ip | net_range | net_fqdn | net_group | service | service_group | zone | urlcat
    name: str
    ip_value: str = ""   # "192.168.1.0/24"
    range_start: str = ""
    range_end: str = ""
    fqdn: str = ""
    protocol: int = 6    # TCP=6, UDP=17, ICMP=1
    dst_port_start: int = 0
    dst_port_end: int = 0
    member_ids: List[str] = []
    urls: List[str] = []  # URL Category entries


class ObjectDeleteRequest(BaseModel):
    ext_ids: List[str]


class ObjectUpdateRequest(BaseModel):
    ext_id: str
    device_group_id: str
    obj_type: str        # same as ObjectCreateRequest
    name: str
    ip_value: str = ""
    range_start: str = ""
    range_end: str = ""
    fqdn: str = ""
    protocol: int = 6
    dst_port_start: int = 0
    dst_port_end: int = 0
    member_ids: List[str] = []
    urls: List[str] = []  # URL Category entries


@router.post("/api/v1/objects/create")
async def create_object_endpoint(request: Request, data: ObjectCreateRequest, db: AsyncSession = Depends(get_db)):
    _require_write(request)
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error", "message": "Unauthorized"}, status_code=401)

    client = NGFWClient(user['host'], verify_ssl=False)
    try:
        await client.login(user['username'], user['password'])
        dg_id = data.device_group_id
        name  = data.name.strip()

        if data.obj_type == 'net_ip':
            res = await client.create_network_object({"name": name, "deviceGroupId": dg_id,
                "value": {"inet": {"inet": data.ip_value}}})
            type_lbl, cat = "Host/Network", "net"

        elif data.obj_type == 'net_range':
            res = await client.create_network_object({"name": name, "deviceGroupId": dg_id,
                "value": {"ipRange": {"start": data.range_start, "end": data.range_end}}})
            type_lbl, cat = "Host/Network", "net"

        elif data.obj_type == 'net_fqdn':
            res = await client.create_network_object({"name": name, "deviceGroupId": dg_id,
                "value": {"fqdn": data.fqdn}})
            type_lbl, cat = "Host/Network", "net"

        elif data.obj_type == 'net_group':
            res = await client.create_network_object_group({"name": name, "deviceGroupId": dg_id,
                "items": data.member_ids})
            type_lbl, cat = "Network Group", "net"

        elif data.obj_type == 'service':
            if data.dst_port_end and data.dst_port_end != data.dst_port_start:
                ports = [{"portRange": {"from": data.dst_port_start, "to": data.dst_port_end}}]
            elif data.dst_port_start:
                ports = [{"singlePort": {"port": data.dst_port_start}}]
            else:
                ports = []
            svc_payload: Dict[str, Any] = {"name": name, "deviceGroupId": dg_id, "protocol": data.protocol}
            if ports:
                svc_payload["dstPorts"] = ports
            res = await client.create_service(svc_payload)
            type_lbl, cat = "Service", "service"

        elif data.obj_type == 'service_group':
            res = await client.create_service_group({"name": name, "deviceGroupId": dg_id,
                "serviceIds": data.member_ids})
            type_lbl, cat = "Service Group", "service"

        elif data.obj_type == 'zone':
            res = await client.create_zone({"name": name, "deviceGroupId": dg_id})
            type_lbl, cat = "Security Zone", "zone"

        elif data.obj_type == 'urlcat':
            urls = [u.strip() for u in data.urls if u.strip()]
            res = await client.create_url_category({"name": name, "deviceGroupId": dg_id, "urls": urls})
            type_lbl, cat = "URL Category", "urlcat"

        else:
            return JSONResponse({"status": "error", "message": f"Unknown obj_type: {data.obj_type}"}, status_code=400)

        # Extract ID from response (API sometimes nests it)
        ext_id = res.get("id")
        if not ext_id:
            for v in res.values():
                if isinstance(v, dict) and "id" in v:
                    ext_id = v["id"]
                    break
        if not ext_id:
            raise RuntimeError(f"API did not return object ID: {res}")

        db.add(CachedObject(ext_id=ext_id, name=name, type=type_lbl, category=cat,
                            device_group_id=dg_id, data=res))
        await _log_change(db, user, "create", "object", name, ext_id, dg_id,
                          f"Type: {type_lbl}")
        await db.commit()
        return JSONResponse({"status": "ok", "id": ext_id, "name": name})
    except Exception as e:
        logger.error(f"Object creation failed: {e}")
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)
    finally:
        await client.close()


@router.post("/api/v1/objects/update")
async def update_object_endpoint(request: Request, data: ObjectUpdateRequest, db: AsyncSession = Depends(get_db)):
    _require_write(request)
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error", "message": "Unauthorized"}, status_code=401)

    client = NGFWClient(user['host'], verify_ssl=False)
    try:
        await client.login(user['username'], user['password'])
        dg_id = data.device_group_id
        name  = data.name.strip()
        oid   = data.ext_id

        if data.obj_type == 'net_ip':
            payload = {"id": oid, "name": name, "deviceGroupId": dg_id,
                       "value": {"inet": {"inet": data.ip_value}}}
            type_lbl = "Host/Network"
        elif data.obj_type == 'net_range':
            payload = {"id": oid, "name": name, "deviceGroupId": dg_id,
                       "value": {"ipRange": {"start": data.range_start, "end": data.range_end}}}
            type_lbl = "Host/Network"
        elif data.obj_type == 'net_fqdn':
            payload = {"id": oid, "name": name, "deviceGroupId": dg_id,
                       "value": {"fqdn": data.fqdn}}
            type_lbl = "Host/Network"
        elif data.obj_type == 'net_group':
            payload = {"id": oid, "name": name, "deviceGroupId": dg_id,
                       "items": data.member_ids}
            type_lbl = "Network Group"
        elif data.obj_type == 'service':
            if data.dst_port_end and data.dst_port_end != data.dst_port_start:
                ports = [{"portRange": {"from": data.dst_port_start, "to": data.dst_port_end}}]
            elif data.dst_port_start:
                ports = [{"singlePort": {"port": data.dst_port_start}}]
            else:
                ports = []
            payload = {"id": oid, "name": name, "deviceGroupId": dg_id, "protocol": data.protocol}
            if ports:
                payload["dstPorts"] = ports
            type_lbl = "Service"
        elif data.obj_type == 'service_group':
            payload = {"id": oid, "name": name, "deviceGroupId": dg_id,
                       "serviceIds": data.member_ids}
            type_lbl = "Service Group"
        elif data.obj_type == 'urlcat':
            urls = [u.strip() for u in data.urls if u.strip()]
            payload = {"id": oid, "name": name, "deviceGroupId": dg_id, "urls": urls}
            type_lbl = "URL Category"
        elif data.obj_type == 'zone':
            payload = {"id": oid, "name": name, "deviceGroupId": dg_id}
            type_lbl = "Security Zone"
        else:
            return JSONResponse({"status": "error", "message": f"Update not supported for type: {data.obj_type}"}, status_code=400)

        ok = await client.update_object(type_lbl, payload)
        if not ok:
            return JSONResponse({"status": "error", "message": "NGFW returned error"}, status_code=500)

        # Update local cache
        obj_res = await db.execute(select(CachedObject).where(CachedObject.ext_id == oid))
        cached = obj_res.scalar_one_or_none()
        if cached:
            cached.name = name
            update_data: Dict[str, Any] = {**(cached.data or {}), "name": name}
            if data.obj_type == 'urlcat':
                update_data["urls"] = [u.strip() for u in data.urls if u.strip()]
            cached.data = update_data

        await _log_change(db, user, "update", "object", name, oid, dg_id, f"Type: {type_lbl}")
        await db.commit()
        return JSONResponse({"status": "ok"})
    except Exception as e:
        logger.error(f"Object update failed: {e}")
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)
    finally:
        await client.close()


@router.post("/api/v1/objects/delete")
async def delete_objects_endpoint(request: Request, data: ObjectDeleteRequest, db: AsyncSession = Depends(get_db)):
    _require_write(request)
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error", "message": "Unauthorized"}, status_code=401)

    stmt = select(CachedObject).where(CachedObject.ext_id.in_(data.ext_ids))
    objects = (await db.execute(stmt)).scalars().all()

    client = NGFWClient(user['host'], verify_ssl=False)
    try:
        await client.login(user['username'], user['password'])
        failed = []
        for obj in objects:
            ok = await client.delete_object(obj.type, obj.ext_id)
            if ok:
                await _log_change(db, user, "delete", "object", obj.name, obj.ext_id,
                                  obj.device_group_id or "")
                await db.delete(obj)
            else:
                failed.append(obj.name)
        await db.commit()
        if failed:
            return JSONResponse({"status": "partial", "failed": failed, "deleted": len(objects) - len(failed)})
        return JSONResponse({"status": "ok", "deleted": len(objects)})
    except Exception as e:
        logger.error(f"Object deletion failed: {e}")
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)
    finally:
        await client.close()


# ===========================================================================
#  OBJECT FIND & REPLACE — Phase 7.7
# ===========================================================================

class ReplaceInRulesRequest(BaseModel):
    old_ext_id: str
    new_ext_id: str
    device_group_id: str


def _ids_from_field(field: dict) -> List[str]:
    """Extract UUID list from a rule field in any NGFW API format."""
    if not field:
        return []
    kind = field.get("kind", "")
    if "ANY" in kind:
        return []
    objects = field.get("objects", [])
    if isinstance(objects, dict):
        return [str(x) for x in objects.get("array", [])]
    ids: List[str] = []
    for obj in objects:
        if not isinstance(obj, dict):
            continue
        if "id" in obj:
            ids.append(obj["id"])
        else:
            for v in obj.values():
                if isinstance(v, dict) and "id" in v:
                    ids.append(v["id"])
                    break
    return ids


def _field_with_ids(field: dict, ids: List[str], any_kind: str, list_kind: str) -> dict:
    """Rebuild a field in simplified array format."""
    if not ids:
        orig_kind = (field or {}).get("kind", any_kind)
        return {"kind": orig_kind if "ANY" in orig_kind else any_kind, "objects": {"array": []}}
    return {"kind": list_kind, "objects": {"array": ids}}


@router.post("/api/v1/objects/replace-in-rules")
async def replace_object_in_rules(
    request: Request, data: ReplaceInRulesRequest, db: AsyncSession = Depends(get_db)
):
    """Replace old_ext_id with new_ext_id in all SEC rules of a device group."""
    _require_write(request)
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error", "message": "Unauthorized"}, status_code=401)

    folders_stmt = select(Folder).where(Folder.device_group_id == data.device_group_id)
    folder_ids = [f.id for f in (await db.execute(folders_stmt)).scalars().all()]
    if not folder_ids:
        return JSONResponse({"status": "ok", "updated": 0, "failed": 0})

    rules = (await db.execute(
        select(CachedRule).where(CachedRule.folder_id.in_(folder_ids))
    )).scalars().all()

    # field_name → (any_kind, list_kind)
    FIELDS = {
        "sourceAddr":      ("RULE_KIND_ANY",      "RULE_KIND_LIST"),
        "destinationAddr": ("RULE_KIND_ANY",       "RULE_KIND_LIST"),
        "service":         ("RULE_KIND_ANY",       "RULE_KIND_LIST"),
        "application":     ("RULE_KIND_ANY",       "RULE_KIND_LIST"),
        "urlCategory":     ("RULE_KIND_ANY",       "RULE_KIND_LIST"),
        "sourceZone":      ("RULE_KIND_ANY",       "RULE_KIND_LIST"),
        "destinationZone": ("RULE_KIND_ANY",       "RULE_KIND_LIST"),
        "sourceUser":      ("RULE_USER_KIND_ANY",  "RULE_USER_KIND_LIST"),
    }

    old_id = data.old_ext_id
    new_id = data.new_ext_id

    # Pre-filter: only rules that actually reference old_id
    affected = []
    for rule in rules:
        d = rule.data or {}
        for fname in FIELDS:
            if old_id in _ids_from_field(d.get(fname)):
                affected.append(rule)
                break

    if not affected:
        return JSONResponse({"status": "ok", "updated": 0, "failed": 0})

    client = NGFWClient(user['host'], verify_ssl=False)
    try:
        await client.login(user['username'], user['password'])
        updated = 0
        failed = 0

        for rule in affected:
            d = rule.data or {}

            def _get_field(fname: str) -> dict:
                f = d.get(fname, {})
                ids = _ids_from_field(f)
                ids = [new_id if x == old_id else x for x in ids]
                ak, lk = FIELDS[fname]
                return _field_with_ids(f, ids, ak, lk)

            api_payload = {
                "id":              rule.ext_id,
                "name":            rule.name,
                "description":     d.get("description", ""),
                "action":          d.get("action", "SECURITY_RULE_ACTION_ALLOW"),
                "enabled":         d.get("enabled", True),
                "logMode":         d.get("logMode", "SECURITY_RULE_LOG_MODE_AT_RULE_HIT"),
                "sourceZone":      _get_field("sourceZone"),
                "destinationZone": _get_field("destinationZone"),
                "sourceAddr":      _get_field("sourceAddr"),
                "destinationAddr": _get_field("destinationAddr"),
                "service":         _get_field("service"),
                "application":     _get_field("application"),
                "urlCategory":     _get_field("urlCategory"),
                "sourceUser":      _get_field("sourceUser"),
            }
            if d.get("ipsProfile"):
                api_payload["ipsProfileId"] = (d["ipsProfile"] or {}).get("id", "")
            if d.get("avProfile"):
                api_payload["avProfileId"] = (d["avProfile"] or {}).get("id", "")

            try:
                await client.update_rule(rule.ext_id, api_payload)
                # Update local cache with simplified format for each replaced field
                for fname in FIELDS:
                    f = d.get(fname, {})
                    ids = _ids_from_field(f)
                    if old_id in ids:
                        ak, lk = FIELDS[fname]
                        new_ids = [new_id if x == old_id else x for x in ids]
                        d[fname] = _field_with_ids(f, new_ids, ak, lk)
                rule.data = d
                await _log_change(
                    db, user, "update", "rule", rule.name, rule.ext_id,
                    data.device_group_id,
                    f"Object replaced: {old_id[:8]}… → {new_id[:8]}…",
                )
                updated += 1
            except Exception as e:
                logger.error(f"Replace failed for rule '{rule.name}': {e}")
                failed += 1

        await db.commit()
        return JSONResponse({"status": "ok", "updated": updated, "failed": failed})
    except Exception as e:
        logger.error(f"Replace-in-rules error: {e}")
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)
    finally:
        await client.close()


# ===========================================================================
#  NAT RULES
# ===========================================================================

SNAT_LABELS = {
    "NAT_SOURCE_TRANSLATION_TYPE_NONE":           "None",
    "NAT_SOURCE_TRANSLATION_TYPE_DYNAMIC_IP_PORT": "PAT",
    "NAT_SOURCE_TRANSLATION_TYPE_STATIC_IP":       "Static IP",
    "NAT_SOURCE_TRANSLATION_TYPE_STATIC_IP_PORT":  "Static IP+Port",
}
DNAT_LABELS = {
    "NAT_DESTINATION_TRANSLATION_TYPE_NONE":         "None",
    "NAT_DESTINATION_TRANSLATION_TYPE_ADDRESS_POOL": "DNAT",
}


def nat_to_dict(rule: CachedNatRule, object_map: Dict[str, Any], device_group_id: str = "") -> Dict[str, Any]:
    d = rule.data or {}

    def _extract(field_key: str):
        section = d.get(field_key)
        if not section:
            return []
        kind = section.get("kind", "")
        if "ANY" in kind:
            return []
        objects = section.get("objects", [])
        if isinstance(objects, dict):
            objects = objects.get("array", [])
        if not isinstance(objects, list):
            return []
        result = []
        for item in objects:
            if not isinstance(item, dict):
                continue
            if "id" in item:
                result.append((item["id"], item.get("name")))
            else:
                for v in item.values():
                    if isinstance(v, dict) and "id" in v:
                        result.append((v["id"], v.get("name")))
                        break
        return result

    def _render(field_key: str, css: str = "") -> str:
        items = _extract(field_key)
        if not items:
            return "Any"
        tags = []
        for uid, emb_name in items[:4]:
            cached = object_map.get(uid)
            name = (cached.name if cached else emb_name) or uid[:8]
            cls = f"obj-tag {css}".strip()
            tags.append(
                f"<span class='{cls}' title='{name}'"
                f" data-ext-id='{uid}' data-dg='{device_group_id}'>{name}</span>"
            )
        extra = len(items) - 4
        result = "".join(tags)
        if extra > 0:
            result += f"<span class='obj-tag overflow'>+{extra}</span>"
        return result

    snat_raw = d.get("srcTranslationType", "")
    dnat_raw = d.get("dstTranslationType", "")

    return {
        "id":            rule.id,
        "ext_id":        rule.ext_id,
        "name":          rule.name,
        "folder_id":     rule.folder_id,
        "device_group_id": rule.device_group_id,
        "description":   d.get("description", ""),
        "enabled":       d.get("enabled", True),
        "snat_type":     SNAT_LABELS.get(snat_raw, snat_raw or "None"),
        "dnat_type":     DNAT_LABELS.get(dnat_raw, dnat_raw or "None"),
        "snat_raw":      snat_raw,
        "dnat_raw":      dnat_raw,
        "src_addr_type": d.get("srcTranslationAddrType", "NAT_SOURCE_TRANSLATION_ADDRESS_TYPE_NONE"),
        "dst_translated_port": d.get("dstTranslatedPort", ""),
        "src_zone":  _render("sourceZone", "zone"),
        "src_net":   _render("sourceAddr"),
        "dst_zone":  _render("destinationZone", "zone"),
        "dst_net":   _render("destinationAddr"),
        "service":   _render("service", "service"),
        "is_modified": rule.is_modified or False,
        "modified_at": rule.modified_at or "",
        # raw IDs for edit modal
        "src_zone_ids":  [uid for uid, _ in _extract("sourceZone")],
        "dst_zone_ids":  [uid for uid, _ in _extract("destinationZone")],
        "src_net_ids":   [uid for uid, _ in _extract("sourceAddr")],
        "dst_net_ids":   [uid for uid, _ in _extract("destinationAddr")],
        "service_ids":   [uid for uid, _ in _extract("service")],
        "src_translated_ids": [uid for uid, _ in _extract("srcTranslatedAddress")] if isinstance(d.get("srcTranslatedAddress"), dict) else [],
        "dst_translated_ids": [uid for uid, _ in _extract("dstTranslatedAddress")] if isinstance(d.get("dstTranslatedAddress"), dict) else [],
    }


class NatCreateRequest(BaseModel):
    folder_id: str
    name: str
    description: str = ""
    enabled: bool = True
    src_zone_ids: List[str] = []
    dst_zone_ids: List[str] = []
    src_net_ids: List[str] = []
    dst_net_ids: List[str] = []
    service_ids: List[str] = []
    snat_type: str = "NAT_SOURCE_TRANSLATION_TYPE_NONE"
    src_addr_type: str = "NAT_SOURCE_TRANSLATION_ADDRESS_TYPE_NONE"
    src_translated_ids: List[str] = []
    dnat_type: str = "NAT_DESTINATION_TRANSLATION_TYPE_NONE"
    dst_translated_ids: List[str] = []
    dst_translated_port: int = 0


class NatDeleteRequest(BaseModel):
    rule_ids: List[str]


class NatReorderRequest(BaseModel):
    folder_id: str
    rule_ids: List[str]


class NatToggleRequest(BaseModel):
    rule_id: str
    enabled: bool


class NatUpdateRequest(BaseModel):
    rule_id: str          # local DB id
    name: str
    description: str = ""
    enabled: bool = True
    src_zone_ids: List[str] = []
    dst_zone_ids: List[str] = []
    src_net_ids: List[str] = []
    dst_net_ids: List[str] = []
    service_ids: List[str] = []
    snat_type: str = "NAT_SOURCE_TRANSLATION_TYPE_NONE"
    src_addr_type: str = "NAT_SOURCE_TRANSLATION_ADDRESS_TYPE_NONE"
    src_translated_ids: List[str] = []
    dnat_type: str = "NAT_DESTINATION_TRANSLATION_TYPE_NONE"
    dst_translated_ids: List[str] = []
    dst_translated_port: int = 0


@router.get("/nat", response_class=HTMLResponse)
async def nat_page(request: Request, folder_id: str = Query(None), db: AsyncSession = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return RedirectResponse(url="/login")
    spa = _try_spa()
    if spa: return spa

    # Show ALL NAT folders from ALL СУ — IDs are scoped, no collision risk
    nat_stmt = select(NatFolder).options(selectinload(NatFolder.rules)).order_by(NatFolder.device_group_id, NatFolder.sort_order)
    all_folders = (await db.execute(nat_stmt)).scalars().all()

    device_names = {o.device_id: o.name for o in (await db.execute(select(DeviceMeta))).scalars().all()}

    tree: Dict[str, Any] = {}
    first_folder_id = None
    first_device_id = None

    for f in all_folders:
        if not first_folder_id:
            first_folder_id = str(f.id)
        if not first_device_id:
            first_device_id = f.device_group_id
        gid = f.device_group_id or "unknown"
        dev_name = device_names.get(gid, f"Device {gid[:8]}")
        if gid not in tree:
            tree[gid] = {"name": dev_name, "id": gid, "sections": {"pre": [], "post": [], "default": []}}
        f.device_name = dev_name
        f.rules_count = len(f.rules)
        sec = f.section.lower() if f.section and f.section.lower() in ['pre', 'post', 'default'] else 'pre'
        tree[gid]["sections"][sec].append(f)

    target_folder = next((f for f in all_folders if str(f.id) == folder_id), None)
    if folder_id and not target_folder:
        folder_id = first_folder_id
        target_folder = next((f for f in all_folders if str(f.id) == folder_id), None)

    selected_fid = folder_id or first_folder_id
    current_device_id = target_folder.device_group_id if target_folder else first_device_id

    nat_obj_stmt = select(CachedObject.ext_id, CachedObject.name, CachedObject.data)

    class MockObj:
        def __init__(self, ext_id, name, data):
            self.ext_id = ext_id; self.name = name; self.data = data

    object_map: Dict[str, Any] = {}
    for row in (await db.execute(nat_obj_stmt)).fetchall():
        obj = MockObj(row.ext_id, row.name, row.data)
        object_map[row.ext_id] = obj
        if ":" in row.ext_id:
            object_map.setdefault(row.ext_id.split(":", 1)[1], obj)

    dashboard_data = []
    if target_folder:
        rules_sorted = sorted(target_folder.rules, key=lambda x: x.folder_sort_order)
        nat_dg_id = target_folder.device_group_id or ""
        dashboard_data.append({
            "folder": target_folder,
            "rules": [nat_to_dict(r, object_map, nat_dg_id) for r in rules_sorted],
        })

    filtered_tree = {k: v for k, v in tree.items() if not k.endswith(":global") and k != "global"}
    su_tree = await _build_su_tree(db, filtered_tree)

    return templates.TemplateResponse(request, "nat.html", base_ctx(request,
        tree=filtered_tree, su_tree=su_tree, dashboard_data=dashboard_data,
        selected_folder_id=selected_fid, current_device_id=current_device_id,
        user=user,
    ))


@router.post("/nat/create_folder")
async def create_nat_folder(
    request: Request,
    folder_name: str = Form(...),
    device_id: str = Form(...),
    section: str = Form(...),
    db: AsyncSession = Depends(get_db),
):
    nat_user = get_current_user(request)
    if not nat_user:
        return RedirectResponse("/login", status_code=303)
    stmt = select(func.max(NatFolder.sort_order)).where(NatFolder.device_group_id == device_id)
    max_sort = (await db.execute(stmt)).scalar() or 0
    new_folder = NatFolder(
        id=str(uuid.uuid4()),
        name=folder_name,
        device_group_id=device_id,
        section=section,
        sort_order=max_sort + 1,
        su_id=nat_user.get('su_id'),
    )
    db.add(new_folder)
    await db.commit()
    return RedirectResponse(url=f"/nat?folder_id={new_folder.id}", status_code=303)


@router.post("/nat/deploy")
async def deploy_nat(request: Request, device_id: str = Form(...), db: AsyncSession = Depends(get_db)):
    _require_write(request)
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error", "message": "Unauthorized"}, status_code=401)
    client = NGFWClient(user['host'], verify_ssl=False)
    try:
        await client.login(user['username'], user['password'])
        from app.services.nat_service import NatDeployService
        await NatDeployService().deploy_nat_policy(db, client, device_id)
        return JSONResponse({"status": "ok"})
    except Exception as e:
        logger.error(f"NAT deploy failed: {e}")
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)
    finally:
        await client.close()


@router.get("/api/v1/nat/folders/tree")
async def nat_folders_tree(device_group_id: str = Query(None), db: AsyncSession = Depends(get_db)):
    stmt = select(NatFolder).where(NatFolder.device_group_id == device_group_id).order_by(NatFolder.section, NatFolder.sort_order)
    folders = (await db.execute(stmt)).scalars().all()
    return JSONResponse([{"id": f.id, "name": f.name, "section": f.section} for f in folders])


@router.post("/api/v1/nat/rules/create")
async def create_nat_rule_endpoint(request: Request, data: NatCreateRequest, db: AsyncSession = Depends(get_db)):
    _require_write(request)
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error", "message": "Unauthorized"}, status_code=401)

    folder = await db.get(NatFolder, data.folder_id)
    if not folder:
        return JSONResponse({"status": "error", "message": "Folder not found"}, status_code=404)

    def build_field(ids):
        if not ids:
            return {"kind": "RULE_KIND_ANY", "objects": {"array": []}}
        return {"kind": "RULE_KIND_LIST", "objects": {"array": list(ids)}}

    section = (folder.section or "pre").lower()

    api_payload: Dict[str, Any] = {
        "name": data.name,
        "description": data.description,
        "deviceGroupId": folder.device_group_id,
        "precedence": section,
        "position": 1,
        "enabled": data.enabled,
        "srcTranslationType": data.snat_type,
        "srcTranslationAddrType": data.src_addr_type,
        "dstTranslationType": data.dnat_type,
        "sourceZone":      build_field(data.src_zone_ids),
        "sourceAddr":      build_field(data.src_net_ids),
        "destinationZone": build_field(data.dst_zone_ids),
        "destinationAddr": build_field(data.dst_net_ids),
        "service":         build_field(data.service_ids),
    }
    if data.src_translated_ids:
        api_payload["srcTranslatedAddress"] = data.src_translated_ids
    if data.dst_translated_ids:
        api_payload["dstTranslatedAddress"] = data.dst_translated_ids
    if data.dst_translated_port:
        api_payload["dstTranslatedPort"] = data.dst_translated_port

    client = NGFWClient(user['host'], verify_ssl=False)
    try:
        await client.login(user['username'], user['password'])
        res = await client.create_nat_rule(api_payload)
        ext_id = res.get("id")
        if not ext_id:
            raise RuntimeError(f"API did not return NAT rule ID: {res}")

        stmt = select(func.max(CachedNatRule.folder_sort_order)).where(CachedNatRule.folder_id == data.folder_id)
        max_pos = (await db.execute(stmt)).scalar() or 0

        nat_entry = CachedNatRule(
            id=str(uuid.uuid4()),
            ext_id=ext_id,
            name=data.name,
            folder_id=data.folder_id,
            folder_sort_order=max_pos + 1,
            device_group_id=folder.device_group_id,
            data={**api_payload, "id": ext_id},
        )
        db.add(nat_entry)
        await _log_change(db, user, "create", "nat_rule", data.name, ext_id,
                          folder.device_group_id)
        await db.commit()
        return JSONResponse({"status": "ok"})
    except Exception as e:
        logger.error(f"NAT rule creation failed: {e}")
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)
    finally:
        await client.close()


@router.post("/api/v1/nat/rules/delete")
async def delete_nat_rules(request: Request, data: NatDeleteRequest, db: AsyncSession = Depends(get_db)):
    _require_write(request)
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error", "message": "Unauthorized"}, status_code=401)
    client = NGFWClient(user['host'], verify_ssl=False)
    try:
        await client.login(user['username'], user['password'])
        stmt = select(CachedNatRule).where(CachedNatRule.id.in_(data.rule_ids))
        rules = (await db.execute(stmt)).scalars().all()
        for rule in rules:
            await client.delete_nat_rule(rule.ext_id)
            await _log_change(db, user, "delete", "nat_rule", rule.name, rule.ext_id,
                              rule.device_group_id or "")
            await db.delete(rule)
        await db.commit()
        return JSONResponse({"status": "ok"})
    except Exception as e:
        logger.error(f"NAT delete failed: {e}")
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)
    finally:
        await client.close()


@router.post("/api/v1/nat/rules/reorder")
async def reorder_nat_rules(data: NatReorderRequest, db: AsyncSession = Depends(get_db)):
    stmt = select(CachedNatRule).where(CachedNatRule.id.in_(data.rule_ids))
    rules = (await db.execute(stmt)).scalars().all()
    rule_map = {r.id: r for r in rules}
    for index, r_id in enumerate(data.rule_ids):
        if r_id in rule_map:
            rule_map[r_id].folder_id = data.folder_id
            rule_map[r_id].folder_sort_order = index
    await db.commit()
    return JSONResponse({"status": "ok"})


@router.post("/api/v1/nat/rules/toggle")
async def toggle_nat_rule(request: Request, data: NatToggleRequest, db: AsyncSession = Depends(get_db)):
    _require_write(request)
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error", "message": "Unauthorized"}, status_code=401)
    rule = await db.get(CachedNatRule, data.rule_id)
    if not rule:
        return JSONResponse({"status": "error", "message": "NAT rule not found"}, status_code=404)
    client = NGFWClient(user['host'], verify_ssl=False)
    try:
        await client.login(user['username'], user['password'])
        await client.update_rule(rule.ext_id, {"id": rule.ext_id, "enabled": data.enabled})
        if rule.data:
            rule.data = {**rule.data, "enabled": data.enabled}
        state = "enabled" if data.enabled else "disabled"
        await _log_change(db, user, "toggle", "nat_rule", rule.name, rule.ext_id,
                          rule.device_group_id or "", f"Set {state}")
        await db.commit()
        return JSONResponse({"status": "ok"})
    except Exception as e:
        logger.error(f"NAT toggle failed: {e}")
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)
    finally:
        await client.close()


@router.post("/api/v1/nat/rules/update")
async def update_nat_rule_endpoint(request: Request, data: NatUpdateRequest, db: AsyncSession = Depends(get_db)):
    _require_write(request)
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error", "message": "Unauthorized"}, status_code=401)

    rule = await db.get(CachedNatRule, data.rule_id)
    if not rule:
        return JSONResponse({"status": "error", "message": "NAT rule not found"}, status_code=404)

    def build_field(ids):
        if not ids:
            return {"kind": "RULE_KIND_ANY", "objects": {"array": []}}
        return {"kind": "RULE_KIND_LIST", "objects": {"array": list(ids)}}

    existing = rule.data or {}
    api_payload: Dict[str, Any] = {
        "id": rule.ext_id,
        "name": data.name,
        "description": data.description,
        "enabled": data.enabled,
        "srcTranslationType": data.snat_type,
        "srcTranslationAddrType": data.src_addr_type,
        "dstTranslationType": data.dnat_type,
        "sourceZone":      build_field(data.src_zone_ids),
        "sourceAddr":      build_field(data.src_net_ids),
        "destinationZone": build_field(data.dst_zone_ids),
        "destinationAddr": build_field(data.dst_net_ids),
        "service":         build_field(data.service_ids),
    }
    if data.src_translated_ids:
        api_payload["srcTranslatedAddress"] = data.src_translated_ids
    if data.dst_translated_ids:
        api_payload["dstTranslatedAddress"] = data.dst_translated_ids
    if data.dst_translated_port:
        api_payload["dstTranslatedPort"] = data.dst_translated_port

    client = NGFWClient(user['host'], verify_ssl=False)
    try:
        await client.login(user['username'], user['password'])
        ok = await client.update_nat_rule(api_payload)
        if not ok:
            return JSONResponse({"status": "error", "message": "NGFW returned error"}, status_code=500)

        rule.name = data.name
        rule.data = {**existing, **api_payload}
        await _log_change(db, user, "update", "nat_rule", data.name, rule.ext_id,
                          rule.device_group_id or "")
        await db.commit()
        return JSONResponse({"status": "ok"})
    except Exception as e:
        logger.error(f"NAT update failed: {e}")
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)
    finally:
        await client.close()





# ===========================================================================
#  LOGS & MONITORING
# ===========================================================================

@router.get("/logs", response_class=HTMLResponse)
async def logs_page(request: Request, su_id: str = Query(None), db: AsyncSession = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return RedirectResponse(url="/login")
    spa = _try_spa()
    if spa: return spa

    su_list_res = await db.execute(select(ManagementSystem).order_by(ManagementSystem.created_at))
    su_list = su_list_res.scalars().all()

    # Resolve selected SU: URL param → session su_id → first available
    session_su_id = user.get('su_id')
    selected_su_id = (
        su_id if su_id and any(s.id == su_id for s in su_list)
        else session_su_id if session_su_id and any(s.id == session_su_id for s in su_list)
        else (su_list[0].id if su_list else None)
    )

    # Find first (main) device of the selected SU — used as device_group_id for log fetching
    meta_res = await db.execute(select(DeviceMeta).where(
        DeviceMeta.su_id == selected_su_id
    ).where(
        ~DeviceMeta.device_id.endswith(":global")
    ).where(
        DeviceMeta.device_id != "global"
    ).order_by(DeviceMeta.name))
    su_devices = meta_res.scalars().all()
    selected_device_id = su_devices[0].device_id if su_devices else ""

    return templates.TemplateResponse(request, "logs.html", base_ctx(request,
        user=user, log_ttl_hours=LOG_TTL_HOURS,
        su_list=su_list, selected_su_id=selected_su_id,
        selected_device_id=selected_device_id,
    ))


# ---- Helpers ----

def _extract_log_fields(row: Dict[str, Any], log_type: str) -> Dict[str, Any]:
    """Pull indexed fields out of a raw NGFW log entry."""
    def _g(*keys):
        for k in keys:
            v = row.get(k)
            if v is not None and v != "":
                return v
        return None

    src_ip   = _g("srcAddr", "sourceAddress", "srcAddress", "srcIp")
    dst_ip   = _g("dstAddr", "dstAddress", "dstIp")
    dst_port = _g("dstPort")
    action   = _g("action", "ruleAction", "verdict")
    ts_raw   = _g("entryGeneration", "entryReceived", "sessionStart", "generateTime", "timestamp")

    event_time = None
    if ts_raw:
        try:
            if isinstance(ts_raw, (int, float)):
                event_time = datetime.fromtimestamp(ts_raw, tz=timezone.utc)
            else:
                s = str(ts_raw).replace("Z", "+00:00")
                event_time = datetime.fromisoformat(s)
                if event_time.tzinfo is None:
                    event_time = event_time.replace(tzinfo=timezone.utc)
        except Exception:
            pass

    try:
        dst_port_int = int(dst_port) if dst_port is not None else None
    except (ValueError, TypeError):
        dst_port_int = None

    return {
        "src_ip":     str(src_ip)[:64]    if src_ip    else None,
        "dst_ip":     str(dst_ip)[:64]    if dst_ip    else None,
        "dst_port":   dst_port_int,
        "action":     str(action)[:64]    if action    else None,
        "event_time": event_time,
    }


def _build_log_query(stmt, log_type: str, src_ip=None, dst_ip=None, dst_port=None,
                     action=None, time_from=None, time_to=None, search=None):
    """Apply SQL filters to a CachedLog SELECT statement."""
    if src_ip:
        stmt = stmt.where(CachedLog.src_ip.ilike(f"%{src_ip}%"))
    if dst_ip:
        stmt = stmt.where(CachedLog.dst_ip.ilike(f"%{dst_ip}%"))
    if dst_port:
        try:
            stmt = stmt.where(CachedLog.dst_port == int(dst_port))
        except ValueError:
            pass
    if action:
        stmt = stmt.where(CachedLog.action.ilike(f"%{action}%"))
    if time_from:
        try:
            tf = datetime.fromisoformat(time_from.replace("Z", "+00:00"))
            if tf.tzinfo is None:
                tf = tf.replace(tzinfo=timezone.utc)
            stmt = stmt.where(CachedLog.event_time >= tf)
        except Exception:
            pass
    if time_to:
        try:
            tt = datetime.fromisoformat(time_to.replace("Z", "+00:00"))
            if tt.tzinfo is None:
                tt = tt.replace(tzinfo=timezone.utc)
            stmt = stmt.where(CachedLog.event_time <= tt)
        except Exception:
            pass
    return stmt


# ---- Pydantic models for log endpoints ----

class LogFetchRequest(BaseModel):
    device_group_id: str
    log_type:        str            # traffic / ips / av / audit
    period_hours:    int = 1        # fallback if time_from/time_to not set
    confirmed_24h:   bool = False   # must be True for spans > 6h
    time_from:       Optional[str] = None  # ISO UTC — overrides period_hours
    time_to:         Optional[str] = None  # ISO UTC
    src_ip:          Optional[str] = None
    dst_ip:          Optional[str] = None
    dst_port:        Optional[int] = None
    action:          Optional[str] = None

class LogBrowseRequest(BaseModel):
    device_group_id: str
    log_type:        str
    limit:           int = 100
    offset:          int = 0
    src_ip:          Optional[str] = None
    dst_ip:          Optional[str] = None
    dst_port:        Optional[int] = None
    action:          Optional[str] = None
    time_from:       Optional[str] = None
    time_to:         Optional[str] = None

class LogClearRequest(BaseModel):
    device_group_id: str
    log_type:        Optional[str] = None  # None = clear all types


# ---- Fetch from NGFW → store in cache ----

@router.post("/api/v1/logs/fetch")
async def fetch_logs(request: Request, data: LogFetchRequest, db: AsyncSession = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error", "message": "Unauthorized"}, status_code=401)

    now_utc = datetime.now(timezone.utc).replace(microsecond=0)

    if data.time_from and data.time_to:
        # Custom time range supplied directly from UI date pickers
        try:
            tf = datetime.fromisoformat(data.time_from.replace("Z", "+00:00"))
            tt = datetime.fromisoformat(data.time_to.replace("Z", "+00:00"))
            if tf.tzinfo is None: tf = tf.replace(tzinfo=timezone.utc)
            if tt.tzinfo is None: tt = tt.replace(tzinfo=timezone.utc)
            if tf >= tt:
                return JSONResponse({"status": "error", "message": "time_from must be before time_to"}, status_code=400)
        except Exception:
            return JSONResponse({"status": "error", "message": "Invalid time_from/time_to format"}, status_code=400)
        time_from = tf.strftime("%Y-%m-%dT%H:%M:%SZ")
        time_to   = tt.strftime("%Y-%m-%dT%H:%M:%SZ")
        hours_span = max(1, (tt - tf).total_seconds() / 3600)
        if hours_span > 6 and not data.confirmed_24h:
            return JSONResponse({
                "status": "warn_24h",
                "message": f"Выгрузка {round(hours_span, 1)}ч логов может занять время. Подтвердите.",
            })
        max_records = min(20_000, max(2_000, int(hours_span * 1_000)))
    else:
        # Fallback: period_hours relative to now
        period = data.period_hours if data.period_hours > 0 else 1
        if period > 6 and not data.confirmed_24h:
            return JSONResponse({
                "status": "warn_24h",
                "message": f"Выгрузка {period}ч логов может занять время. Подтвердите.",
            })
        time_to   = now_utc.strftime("%Y-%m-%dT%H:%M:%SZ")
        # +12h lookback buffer: NGFW collector may have indexing lag
        time_from = (now_utc - timedelta(hours=period + 12)).strftime("%Y-%m-%dT%H:%M:%SZ")
        max_records = min(20_000, max(2_000, period * 1_000))

    extra_filters: Dict[str, Any] = {}
    if data.src_ip:   extra_filters["srcAddr"]  = data.src_ip
    if data.dst_ip:   extra_filters["dstAddr"]  = data.dst_ip
    if data.dst_port: extra_filters["dstPort"]  = data.dst_port
    if data.action:   extra_filters["action"]   = data.action

    # Derive the correct NGFW host from device_group_id (format: {su_id}:{raw_gid})
    _log_su_id = data.device_group_id.split(":")[0] if ":" in data.device_group_id else None
    _log_su = (await db.get(ManagementSystem, _log_su_id)) if _log_su_id else None
    if _log_su:
        _api_host = _log_su.host
        _api_user = _log_su.username
        _api_pass = decrypt_pw(_log_su.password_enc)
        _api_ssl  = getattr(_log_su, 'verify_ssl', False)
    else:
        _api_host = user['host']
        _api_user = user['username']
        _api_pass = user['password']
        _api_ssl  = False

    # Unscope device_group_id before passing to NGFW API
    _raw_gid = data.device_group_id.split(":", 1)[1] if ":" in data.device_group_id else data.device_group_id

    client = NGFWClient(_api_host, verify_ssl=_api_ssl)
    try:
        await client.login(_api_user, _api_pass)
        raw_logs = await client.fetch_all_logs(
            log_type        = data.log_type,
            device_group_id = _raw_gid,
            time_from       = time_from,
            time_to         = time_to,
            extra_filters   = extra_filters,
            max_records     = max_records,
        )
    except Exception as e:
        logger.error(f"fetch_all_logs failed: {e}")
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)
    finally:
        await client.close()

    # Clear only the overlapping time range so previously fetched data outside
    # this window is preserved (e.g. don't wipe 24h cache when re-fetching 1h).
    tf_dt = datetime.fromisoformat(time_from.replace("Z", "+00:00"))
    tt_dt = datetime.fromisoformat(time_to.replace("Z",   "+00:00"))
    del_stmt = (
        sa_delete(CachedLog)
        .where(CachedLog.device_group_id == data.device_group_id)
        .where(CachedLog.log_type == data.log_type)
        .where(
            (CachedLog.event_time >= tf_dt) | CachedLog.event_time.is_(None)
        )
        .where(
            (CachedLog.event_time <= tt_dt) | CachedLog.event_time.is_(None)
        )
    )
    await db.execute(del_stmt)

    fetched_at = datetime.now(timezone.utc)
    batch = []
    for row in raw_logs:
        fields = _extract_log_fields(row, data.log_type)
        batch.append(CachedLog(
            device_group_id = data.device_group_id,
            log_type        = data.log_type,
            event_time      = fields["event_time"],
            src_ip          = fields["src_ip"],
            dst_ip          = fields["dst_ip"],
            dst_port        = fields["dst_port"],
            action          = fields["action"],
            data            = row,
            fetched_at      = fetched_at,
        ))

    if batch:
        db.add_all(batch)
    await db.commit()

    logger.info(f"Fetched and cached {len(batch)} {data.log_type} logs for {data.device_group_id}")
    return JSONResponse({
        "status":     "ok",
        "fetched":    len(raw_logs),
        "stored":     len(batch),
        "period_hours": data.period_hours,
        "fetched_at": fetched_at.isoformat(),
    })


# ---- Query from cache ----

@router.post("/api/v1/logs/query")
async def query_logs(request: Request, data: LogBrowseRequest, db: AsyncSession = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error", "message": "Unauthorized"}, status_code=401)

    base = (
        select(CachedLog)
        .where(CachedLog.device_group_id == data.device_group_id)
        .where(CachedLog.log_type == data.log_type)
    )
    base = _build_log_query(base, data.log_type, data.src_ip, data.dst_ip,
                            data.dst_port, data.action, data.time_from, data.time_to)

    # total count
    count_stmt = select(func.count()).select_from(base.subquery())
    total = (await db.execute(count_stmt)).scalar_one()

    # paginated items
    rows_stmt = (
        base
        .order_by(CachedLog.event_time.desc().nullslast(), CachedLog.id.desc())
        .limit(data.limit)
        .offset(data.offset)
    )
    rows = (await db.execute(rows_stmt)).scalars().all()
    if rows:
        logger.info(f"[query] First row data sample: {str(rows[0].data)[:500]}")
        logger.info(f"[query] First row data type: {type(rows[0].data)}")

    # Cache meta
    meta_stmt = (
        select(func.count(), func.max(CachedLog.fetched_at),
               func.min(CachedLog.event_time), func.max(CachedLog.event_time))
        .where(CachedLog.device_group_id == data.device_group_id)
        .where(CachedLog.log_type == data.log_type)
    )
    meta = (await db.execute(meta_stmt)).one()
    cache_count, last_fetched, oldest_ev, newest_ev = meta[0], meta[1], meta[2], meta[3]

    purge_in_sec = None
    if last_fetched:
        purge_at = last_fetched.replace(tzinfo=timezone.utc) + timedelta(hours=LOG_TTL_HOURS)
        purge_in_sec = max(0, int((purge_at - datetime.now(timezone.utc)).total_seconds()))

    return JSONResponse({
        "status":        "ok",
        "items":         [r.data for r in rows],
        "total":         total,
        "cache_count":   cache_count,
        "fetched_at":    last_fetched.isoformat() if last_fetched else None,
        "purge_in_sec":  purge_in_sec,
        "oldest_event":  oldest_ev.isoformat() if oldest_ev else None,
        "newest_event":  newest_ev.isoformat() if newest_ev else None,
    })


# ---- Status ----

@router.get("/api/v1/logs/status")
async def logs_status(request: Request, device_group_id: str = Query(...), db: AsyncSession = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error"}, status_code=401)

    result = {}
    for lt in ("traffic", "ips", "av", "audit"):
        stmt = (
            select(func.count(), func.max(CachedLog.fetched_at), func.min(CachedLog.event_time), func.max(CachedLog.event_time))
            .where(CachedLog.device_group_id == device_group_id)
            .where(CachedLog.log_type == lt)
        )
        row = (await db.execute(stmt)).one()
        count, last_fetch, oldest_ev, newest_ev = row
        purge_in = None
        if last_fetch:
            purge_at = last_fetch.replace(tzinfo=timezone.utc) + timedelta(hours=LOG_TTL_HOURS)
            purge_in = max(0, int((purge_at - datetime.now(timezone.utc)).total_seconds()))
        result[lt] = {
            "count":       count,
            "fetched_at":  last_fetch.isoformat() if last_fetch else None,
            "oldest_event": oldest_ev.isoformat() if oldest_ev else None,
            "newest_event": newest_ev.isoformat() if newest_ev else None,
            "purge_in_sec": purge_in,
        }
    return JSONResponse({"status": "ok", "types": result})


# ---- Manual clear ----

@router.post("/api/v1/logs/clear")
async def clear_logs(request: Request, data: LogClearRequest, db: AsyncSession = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error"}, status_code=401)

    stmt = sa_delete(CachedLog).where(CachedLog.device_group_id == data.device_group_id)
    if data.log_type:
        stmt = stmt.where(CachedLog.log_type == data.log_type)
    result = await db.execute(stmt)
    await db.commit()
    return JSONResponse({"status": "ok", "deleted": result.rowcount})


# ---- Export CSV (server-side streaming) ----

@router.get("/api/v1/logs/export")
async def export_logs_csv(
    request: Request,
    device_group_id: str = Query(...),
    log_type:        str = Query(...),
    src_ip:          Optional[str] = Query(None),
    dst_ip:          Optional[str] = Query(None),
    dst_port:        Optional[int] = Query(None),
    action:          Optional[str] = Query(None),
    time_from:       Optional[str] = Query(None),
    time_to:         Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
):
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error"}, status_code=401)

    stmt = (
        select(CachedLog)
        .where(CachedLog.device_group_id == device_group_id)
        .where(CachedLog.log_type == log_type)
        .order_by(CachedLog.event_time.desc().nullslast(), CachedLog.id.desc())
    )
    stmt = _build_log_query(stmt, log_type, src_ip, dst_ip, dst_port, action, time_from, time_to)
    rows = (await db.execute(stmt)).scalars().all()

    # Column definitions per log type
    COL_KEYS = {
        "traffic": [("Time","entryGeneration","entryReceived","sessionStart"),
                    ("Context","contextName","deviceName"),
                    ("Src IP","srcAddr"),("Src Port","srcPort"),
                    ("Dst IP","dstAddr"),("Dst Port","dstPort"),
                    ("Protocol","ipProtocol"),("App","app"),
                    ("Action","action"),("Rule","securityRuleName"),
                    ("Bytes Sent","bytesSent"),("Bytes Recv","bytesReceived")],
        "ips":     [("Time","entryGeneration","entryReceived"),
                    ("Context","contextName","deviceName"),
                    ("Src IP","srcAddr"),("Src Port","srcPort"),
                    ("Dst IP","dstAddr"),("Dst Port","dstPort"),
                    ("Threat","threatName"),("Severity","threatSeverity"),
                    ("Action","action"),("Protocol","ipProtocol")],
        "av":      [("Time","entryGeneration","entryReceived"),
                    ("Context","contextName","deviceName"),
                    ("Src IP","srcAddr"),("Dst IP","dstAddr"),
                    ("Threat","threatName"),("File","fileName"),("Action","action")],
        "audit":   [("Time","generateTime"),
                    ("Admin","adminDisplayName","adminLogin"),
                    ("Action","action"),("Source IP","sourceAddress"),
                    ("Args","queryArgs"),("Result","result")],
    }
    cols = COL_KEYS.get(log_type, [("Time","entryGeneration"),("Src","srcAddr"),("Dst","dstAddr"),("Action","action")])

    def _g(d, *keys):
        for k in keys:
            v = d.get(k)
            if v is not None:
                return str(v)
        return ""

    def _generate():
        buf = io.StringIO()
        buf.write("﻿")  # BOM for Excel
        writer = csv.writer(buf)
        writer.writerow([c[0] for c in cols])
        yield buf.getvalue()

        for r in rows:
            buf = io.StringIO()
            writer = csv.writer(buf)
            d = r.data or {}
            writer.writerow([_g(d, *c[1:]) for c in cols])
            yield buf.getvalue()

    filename = f"logs_{log_type}_{device_group_id[:8]}_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    return StreamingResponse(
        _generate(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/api/v1/logs/top_stats")
async def logs_top_stats(
    request: Request,
    device_group_id: str = Query(...),
    log_type: str = Query("traffic"),
    time_from: Optional[str] = Query(None),
    time_to: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
):
    """Aggregate top-N stats from cached logs for dashboard charts."""
    if not get_current_user(request):
        return JSONResponse({"status": "error"}, status_code=401)

    conds = [
        CachedLog.device_group_id == device_group_id,
        CachedLog.log_type == log_type,
    ]
    if time_from:
        try:
            tf = datetime.fromisoformat(time_from.replace("Z", "+00:00"))
            conds.append(CachedLog.event_time >= tf)
        except Exception:
            pass
    if time_to:
        try:
            tt = datetime.fromisoformat(time_to.replace("Z", "+00:00"))
            conds.append(CachedLog.event_time <= tt)
        except Exception:
            pass

    TOP = 10

    async def _top_col(col):
        stmt = (
            select(col, func.count().label("c"))
            .where(*conds)
            .where(col.isnot(None))
            .group_by(col)
            .order_by(func.count().desc())
            .limit(TOP)
        )
        return [{"label": str(r[0]), "count": r[1]} for r in (await db.execute(stmt)).all()]

    # Rule name lives in the JSON data blob
    rule_key = "$.securityRuleName"
    rule_expr = func.json_extract(CachedLog.data, rule_key)
    top_rules_rows = (await db.execute(
        select(rule_expr.label("rn"), func.count().label("c"))
        .where(*conds)
        .where(rule_expr.isnot(None))
        .group_by(rule_expr)
        .order_by(func.count().desc())
        .limit(TOP)
    )).all()

    total = (await db.execute(select(func.count()).where(*conds))).scalar_one()

    return JSONResponse({
        "status":       "ok",
        "total":        total,
        "top_src_ip":   await _top_col(CachedLog.src_ip),
        "top_dst_ip":   await _top_col(CachedLog.dst_ip),
        "top_dst_port": await _top_col(CachedLog.dst_port),
        "top_rules":    [{"label": str(r[0]), "count": r[1]} for r in top_rules_rows],
        "actions":      await _top_col(CachedLog.action),
    })


@router.get("/api/v1/logs/rule_stats")
async def get_rule_stats_endpoint(request: Request, device_group_id: str = Query(...)):
    user = get_current_user(request)
    if not user:
        return JSONResponse([], status_code=401)
    client = NGFWClient(user['host'], verify_ssl=False)
    try:
        await client.login(user['username'], user['password'])
        stats = await client.get_rule_stats(device_group_id)
        return JSONResponse(stats)
    except Exception as e:
        logger.error(f"Rule stats failed: {e}")
        return JSONResponse([])
    finally:
        await client.close()


# =====================================================================
# BLOCK 5 — Policy Rules (Decryption / Authentication / PBR)
# =====================================================================

class PolicyListRequest(BaseModel):
    device_group_id: str
    tab: str  # "decryption" | "auth" | "pbr"

class PolicyCreateRequest(BaseModel):
    device_group_id: str
    tab: str
    payload: Dict[str, Any]

class PolicyDeleteRequest(BaseModel):
    device_group_id: str
    tab: str
    ids: List[str]

class PolicyToggleRequest(BaseModel):
    device_group_id: str
    tab: str
    rule_id: str
    enabled: bool


@router.get("/policy", response_class=HTMLResponse)
async def policy_page(
    request: Request,
    device_id: str = Query(None),
    tab: str = Query("decryption"),
    db: AsyncSession = Depends(get_db),
):
    user = get_current_user(request)
    if not user:
        return RedirectResponse(url="/login")
    spa = _try_spa()
    if spa: return spa

    meta_res = await db.execute(select(DeviceMeta).order_by(DeviceMeta.name))
    all_devices = meta_res.scalars().all()
    devices = [d for d in all_devices if d.device_id != "global"]

    selected_device_id = device_id or (devices[0].device_id if devices else None)

    obj_res = await db.execute(select(CachedObject))
    all_objs = obj_res.scalars().all()

    def _filter_objs(types):
        return [
            {"id": o.ext_id, "name": o.name}
            for o in all_objs
            if o.device_group_id in (selected_device_id, "global") and o.type in types
        ]

    net_objects  = _filter_objs({"Host/Network", "Network", "Network Group"})
    svc_objects  = _filter_objs({"Service", "Service Group"})
    zone_objects = _filter_objs({"Security Zone", "Zone"})

    return templates.TemplateResponse(request, "policy.html", base_ctx(request,
        devices=devices, selected_device_id=selected_device_id,
        active_tab=tab, user=user,
        net_objects=net_objects, svc_objects=svc_objects, zone_objects=zone_objects,
    ))


@router.post("/api/v1/policy/list")
async def list_policy_rules(request: Request, data: PolicyListRequest):
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error", "message": "Unauthorized"}, status_code=401)
    client = NGFWClient(user['host'], verify_ssl=False)
    try:
        await client.login(user['username'], user['password'])
        if data.tab == "decryption":
            rules = await client.list_decryption_rules(data.device_group_id)
        elif data.tab == "auth":
            rules = await client.list_auth_rules(data.device_group_id)
        elif data.tab == "pbr":
            rules = await client.list_pbr_rules(data.device_group_id)
        else:
            return JSONResponse({"status": "error", "message": f"Unknown tab: {data.tab}"}, status_code=400)
        return JSONResponse({"status": "ok", "rules": rules})
    except Exception as e:
        logger.error(f"Policy list ({data.tab}) failed: {e}")
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)
    finally:
        await client.close()


@router.post("/api/v1/policy/create")
async def create_policy_rule(request: Request, data: PolicyCreateRequest, db: AsyncSession = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error", "message": "Unauthorized"}, status_code=401)
    client = NGFWClient(user['host'], verify_ssl=False)
    try:
        await client.login(user['username'], user['password'])
        if data.tab == "decryption":
            result = await client.create_decryption_rule(data.payload)
        elif data.tab == "auth":
            result = await client.create_auth_rule(data.payload)
        elif data.tab == "pbr":
            result = await client.create_pbr_rule(data.payload)
        else:
            return JSONResponse({"status": "error", "message": f"Unknown tab: {data.tab}"}, status_code=400)
        name = data.payload.get("name", "")
        ext_id = (result or {}).get("id", "")
        await _log_change(db, user, "create", f"policy_{data.tab}", name, ext_id,
                          data.device_group_id, f"Tab: {data.tab}")
        await db.commit()
        return JSONResponse({"status": "ok", "rule": result})
    except Exception as e:
        logger.error(f"Policy create ({data.tab}) failed: {e}")
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)
    finally:
        await client.close()


@router.post("/api/v1/policy/update")
async def update_policy_rule(request: Request, data: PolicyCreateRequest, db: AsyncSession = Depends(get_db)):
    """Update a policy rule (same as create but payload must contain 'id')."""
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error", "message": "Unauthorized"}, status_code=401)
    if not data.payload.get("id"):
        return JSONResponse({"status": "error", "message": "payload.id required"}, status_code=400)
    client = NGFWClient(user['host'], verify_ssl=False)
    try:
        await client.login(user['username'], user['password'])
        endpoint_map = {
            "decryption": "UpdateDecryptionRule",
            "auth":       "UpdateAuthenticationRule",
            "pbr":        "UpdatePBRRule",
        }
        endpoint = endpoint_map.get(data.tab)
        if not endpoint:
            return JSONResponse({"status": "error", "message": f"Unknown tab: {data.tab}"}, status_code=400)
        result = await client._create_rule_generic(endpoint, data.payload)
        name = data.payload.get("name", "")
        await _log_change(db, user, "update", f"policy_{data.tab}", name,
                          data.payload["id"], data.device_group_id, f"Tab: {data.tab}")
        await db.commit()
        return JSONResponse({"status": "ok", "rule": result})
    except Exception as e:
        logger.error(f"Policy update ({data.tab}) failed: {e}")
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)
    finally:
        await client.close()


@router.post("/api/v1/policy/delete")
async def delete_policy_rules(request: Request, data: PolicyDeleteRequest, db: AsyncSession = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error", "message": "Unauthorized"}, status_code=401)
    client = NGFWClient(user['host'], verify_ssl=False)
    try:
        await client.login(user['username'], user['password'])
        failed = []
        for rule_id in data.ids:
            if data.tab == "decryption":
                ok = await client.delete_decryption_rule(rule_id)
            elif data.tab == "auth":
                ok = await client.delete_auth_rule(rule_id)
            elif data.tab == "pbr":
                ok = await client.delete_pbr_rule(rule_id)
            else:
                ok = False
            if ok:
                await _log_change(db, user, "delete", f"policy_{data.tab}", rule_id, rule_id,
                                  data.device_group_id, f"Tab: {data.tab}")
            else:
                failed.append(rule_id)
        await db.commit()
        if failed:
            return JSONResponse({"status": "partial", "failed": failed})
        return JSONResponse({"status": "ok", "deleted": len(data.ids)})
    except Exception as e:
        logger.error(f"Policy delete ({data.tab}) failed: {e}")
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)
    finally:
        await client.close()


@router.post("/api/v1/policy/toggle")
async def toggle_policy_rule(request: Request, data: PolicyToggleRequest):
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error", "message": "Unauthorized"}, status_code=401)
    client = NGFWClient(user['host'], verify_ssl=False)
    try:
        await client.login(user['username'], user['password'])
        if data.tab == "decryption":
            ok = await client.toggle_decryption_rule(data.rule_id, data.enabled)
        elif data.tab == "auth":
            ok = await client.toggle_auth_rule(data.rule_id, data.enabled)
        elif data.tab == "pbr":
            ok = await client.toggle_pbr_rule(data.rule_id, data.enabled)
        else:
            ok = False
        return JSONResponse({"status": "ok" if ok else "error"})
    except Exception as e:
        logger.error(f"Policy toggle ({data.tab}) failed: {e}")
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)
    finally:
        await client.close()


# =====================================================================
# BLOCK 6 — System Management (Admins / Backup / Routing / Interfaces / Settings)
# =====================================================================

class AdminCreateRequest(BaseModel):
    device_group_id: str
    login: str
    name: str
    password: str
    role: str = "ReadOnly"

class AdminActionRequest(BaseModel):
    device_group_id: str
    admin_id: str
    action: str  # "delete" | "block" | "unblock"

class AdminPasswordRequest(BaseModel):
    device_group_id: str
    admin_id: str
    new_password: str

class BackupCreateRequest(BaseModel):
    device_group_id: str
    description: str = ""

class BackupDeleteRequest(BaseModel):
    device_group_id: str
    backup_id: str

class SnapshotCommitRequest(BaseModel):
    device_group_id: str
    description: str = ""

class RouteCreateRequest(BaseModel):
    device_group_id: str
    destination: str
    prefix_len: int
    gateway: str
    interface: str = ""
    metric: int = 1
    description: str = ""

class RouteDeleteRequest(BaseModel):
    device_group_id: str
    route_id: str

class TimeoutsSetRequest(BaseModel):
    device_group_id: str
    tcp: Optional[int] = None
    udp: Optional[int] = None
    icmp: Optional[int] = None
    tcp_half_open: Optional[int] = None
    tcp_time_wait: Optional[int] = None
    udp_stream: Optional[int] = None


def _sys_client_from_request(request: Request):
    user = get_current_user(request)
    if not user:
        return None, None
    return user, NGFWClient(user['host'], verify_ssl=False)


@router.get("/system", response_class=HTMLResponse)
async def system_page(
    request: Request,
    su_id: str = Query(None),
    device_id: str = Query(None),
    tab: str = Query("connections"),
    db: AsyncSession = Depends(get_db),
):
    user = get_current_user(request)
    if not user:
        return RedirectResponse(url="/login")
    spa = _try_spa()
    if spa: return spa

    su_res = await db.execute(select(ManagementSystem).order_by(ManagementSystem.created_at))
    su_orm = su_res.scalars().all()
    su_list = [
        {
            "id": s.id, "name": s.name, "host": s.host,
            "username": s.username, "is_active": s.is_active,
            "last_synced_at": s.last_synced_at.isoformat() if s.last_synced_at else None,
        }
        for s in su_orm
    ]

    # Derive su_id from scoped device_id prefix when not explicitly provided
    if not su_id and device_id and ":" in device_id:
        su_id = device_id.split(":")[0]

    session_su_id = user.get('su_id')
    selected_su_id = (
        su_id if su_id and any(s["id"] == su_id for s in su_list)
        else session_su_id if session_su_id and any(s["id"] == session_su_id for s in su_list)
        else (su_list[0]["id"] if su_list else None)
    )

    meta_res = await db.execute(select(DeviceMeta).order_by(DeviceMeta.name))
    all_devices = [d for d in meta_res.scalars().all()
                   if not d.device_id.endswith(":global") and d.device_id != "global"]
    devices = [d for d in all_devices if not selected_su_id or d.su_id == selected_su_id]
    selected_device_id = (
        device_id if device_id and any(d.device_id == device_id for d in devices)
        else (devices[0].device_id if devices else None)
    )

    return templates.TemplateResponse(request, "system.html", base_ctx(request,
        devices=devices, selected_device_id=selected_device_id,
        active_tab=tab, user=user, su_list=su_list, selected_su_id=selected_su_id,
    ))


# ---- Admins ----

@router.get("/api/v1/system/admins")
async def list_admins_endpoint(request: Request, device_group_id: str = Query(...)):
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error"}, status_code=401)
    client = NGFWClient(user['host'], verify_ssl=False)
    try:
        await client.login(user['username'], user['password'])
        admins = await client.list_admins(device_group_id)
        return JSONResponse({"status": "ok", "admins": admins})
    except Exception as e:
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)
    finally:
        await client.close()


@router.post("/api/v1/system/admins/create")
async def create_admin_endpoint(request: Request, data: AdminCreateRequest):
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error"}, status_code=401)
    client = NGFWClient(user['host'], verify_ssl=False)
    try:
        await client.login(user['username'], user['password'])
        payload = {
            "login": data.login, "name": data.name,
            "password": data.password, "role": data.role,
            "deviceGroupId": data.device_group_id,
        }
        result = await client.create_admin(payload)
        return JSONResponse({"status": "ok", "admin": result})
    except Exception as e:
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)
    finally:
        await client.close()


@router.post("/api/v1/system/admins/action")
async def admin_action_endpoint(request: Request, data: AdminActionRequest):
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error"}, status_code=401)
    client = NGFWClient(user['host'], verify_ssl=False)
    try:
        await client.login(user['username'], user['password'])
        if data.action == "delete":
            ok = await client.delete_admin(data.admin_id)
        elif data.action == "block":
            ok = await client.block_admin(data.admin_id)
        elif data.action == "unblock":
            ok = await client.unblock_admin(data.admin_id)
        else:
            return JSONResponse({"status": "error", "message": f"Unknown action: {data.action}"}, status_code=400)
        return JSONResponse({"status": "ok" if ok else "error"})
    except Exception as e:
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)
    finally:
        await client.close()


@router.post("/api/v1/system/admins/password")
async def admin_password_endpoint(request: Request, data: AdminPasswordRequest):
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error"}, status_code=401)
    client = NGFWClient(user['host'], verify_ssl=False)
    try:
        await client.login(user['username'], user['password'])
        ok = await client.update_admin_credentials(data.admin_id, {"password": data.new_password})
        return JSONResponse({"status": "ok" if ok else "error"})
    except Exception as e:
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)
    finally:
        await client.close()


# ---- Backup & Snapshot ----

@router.get("/api/v1/system/backups")
async def list_backups_endpoint(request: Request, device_group_id: str = Query(...)):
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error"}, status_code=401)
    client = NGFWClient(user['host'], verify_ssl=False)
    try:
        await client.login(user['username'], user['password'])
        backups = await client.list_backups(device_group_id)
        snapshots = await client.list_snapshots(device_group_id)
        return JSONResponse({"status": "ok", "backups": backups, "snapshots": snapshots})
    except Exception as e:
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)
    finally:
        await client.close()


@router.post("/api/v1/system/backups/create")
async def create_backup_endpoint(request: Request, data: BackupCreateRequest):
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error"}, status_code=401)
    client = NGFWClient(user['host'], verify_ssl=False)
    try:
        await client.login(user['username'], user['password'])
        result = await client.create_backup(data.device_group_id, data.description)
        return JSONResponse({"status": "ok", "backup": result})
    except Exception as e:
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)
    finally:
        await client.close()


@router.post("/api/v1/system/backups/delete")
async def delete_backup_endpoint(request: Request, data: BackupDeleteRequest):
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error"}, status_code=401)
    client = NGFWClient(user['host'], verify_ssl=False)
    try:
        await client.login(user['username'], user['password'])
        ok = await client.delete_backup(data.backup_id)
        return JSONResponse({"status": "ok" if ok else "error"})
    except Exception as e:
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)
    finally:
        await client.close()


@router.post("/api/v1/system/snapshots/commit")
async def commit_snapshot_endpoint(request: Request, data: SnapshotCommitRequest):
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error"}, status_code=401)
    client = NGFWClient(user['host'], verify_ssl=False)
    try:
        await client.login(user['username'], user['password'])
        result = await client.commit_snapshot(data.device_group_id, data.description)
        return JSONResponse({"status": "ok", "snapshot": result})
    except Exception as e:
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)
    finally:
        await client.close()


# ---- Routing ----

@router.get("/api/v1/system/routing")
async def get_routing_endpoint(request: Request, device_group_id: str = Query(...)):
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error"}, status_code=401)
    client = NGFWClient(user['host'], verify_ssl=False)
    try:
        await client.login(user['username'], user['password'])
        routes, bgp_info, bgp_peers, ospf_info, ospf_areas = await asyncio.gather(
            client.list_static_routes(device_group_id),
            client.get_bgp(device_group_id),
            client.list_bgp_peers(device_group_id),
            client.get_ospf(device_group_id),
            client.list_ospf_areas(device_group_id),
            return_exceptions=True,
        )
        return JSONResponse({
            "status": "ok",
            "routes":     routes     if isinstance(routes, list)     else [],
            "bgp_info":   bgp_info   if isinstance(bgp_info, dict)   else {},
            "bgp_peers":  bgp_peers  if isinstance(bgp_peers, list)  else [],
            "ospf_info":  ospf_info  if isinstance(ospf_info, dict)  else {},
            "ospf_areas": ospf_areas if isinstance(ospf_areas, list) else [],
        })
    except Exception as e:
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)
    finally:
        await client.close()


@router.post("/api/v1/system/routing/create")
async def create_route_endpoint(request: Request, data: RouteCreateRequest):
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error"}, status_code=401)
    client = NGFWClient(user['host'], verify_ssl=False)
    try:
        await client.login(user['username'], user['password'])
        payload = {
            "deviceGroupId": data.device_group_id,
            "destination": data.destination,
            "prefixLen": data.prefix_len,
            "gateway": data.gateway,
            "metric": data.metric,
        }
        if data.interface:
            payload["interface"] = data.interface
        if data.description:
            payload["description"] = data.description
        result = await client.create_static_route(payload)
        return JSONResponse({"status": "ok", "route": result})
    except Exception as e:
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)
    finally:
        await client.close()


@router.post("/api/v1/system/routing/delete")
async def delete_route_endpoint(request: Request, data: RouteDeleteRequest):
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error"}, status_code=401)
    client = NGFWClient(user['host'], verify_ssl=False)
    try:
        await client.login(user['username'], user['password'])
        ok = await client.delete_static_route(data.route_id)
        return JSONResponse({"status": "ok" if ok else "error"})
    except Exception as e:
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)
    finally:
        await client.close()


# ---- Interfaces ----

@router.get("/api/v1/system/interfaces")
async def list_interfaces_endpoint(request: Request, device_group_id: str = Query(...)):
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error"}, status_code=401)
    client = NGFWClient(user['host'], verify_ssl=False)
    try:
        await client.login(user['username'], user['password'])
        virtual, logical = await asyncio.gather(
            client.list_virtual_interfaces(device_group_id),
            client.list_logical_interfaces(device_group_id),
            return_exceptions=True,
        )
        vlist = virtual if isinstance(virtual, list) else []
        llist = logical if isinstance(logical, list) else []
        for i in vlist: i['_itype'] = 'Virtual'
        for i in llist: i['_itype'] = 'Logical'
        return JSONResponse({"status": "ok", "interfaces": vlist + llist})
    except Exception as e:
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)
    finally:
        await client.close()


# ---- Timeouts / Settings ----

@router.get("/api/v1/system/timeouts")
async def get_timeouts_endpoint(request: Request, device_group_id: str = Query(...)):
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error"}, status_code=401)
    client = NGFWClient(user['host'], verify_ssl=False)
    try:
        await client.login(user['username'], user['password'])
        data = await client.get_device_timeouts(device_group_id)
        return JSONResponse({"status": "ok", "timeouts": data})
    except Exception as e:
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)
    finally:
        await client.close()


@router.post("/api/v1/system/timeouts/set")
async def set_timeouts_endpoint(request: Request, data: TimeoutsSetRequest):
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error"}, status_code=401)
    client = NGFWClient(user['host'], verify_ssl=False)
    try:
        await client.login(user['username'], user['password'])
        payload = {k: v for k, v in {
            "tcp": data.tcp, "udp": data.udp, "icmp": data.icmp,
            "tcpHalfOpen": data.tcp_half_open,
            "tcpTimeWait": data.tcp_time_wait,
            "udpStream":   data.udp_stream,
        }.items() if v is not None}
        ok = await client.set_device_timeouts(data.device_group_id, payload)
        return JSONResponse({"status": "ok" if ok else "error"})
    except Exception as e:
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)
    finally:
        await client.close()


# ---------------------------------------------------------------------------
# Policy Analyzer
# ---------------------------------------------------------------------------

class AnalyzerRequest(BaseModel):
    device_group_id: str
    folder_id: Optional[str] = None


@router.get("/analyzer", response_class=HTMLResponse)
async def analyzer_page(request: Request, su_id: str = Query(None), db: AsyncSession = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login")
    spa = _try_spa()
    if spa: return spa

    import app.state as _state
    current_su_id = su_id or user.get('su_id') or (
        _state.su_list[0]["id"] if _state.su_list else None
    )
    meta_stmt = select(DeviceMeta)
    if current_su_id:
        meta_stmt = meta_stmt.where(DeviceMeta.su_id == current_su_id)
    meta_res = await db.execute(meta_stmt)
    device_names = {o.device_id: o.name for o in meta_res.scalars().all()}

    stmt = select(Folder).order_by(Folder.device_group_id, Folder.sort_order)
    if current_su_id:
        stmt = stmt.where(Folder.su_id == current_su_id)
    result = await db.execute(stmt)
    all_folders = result.scalars().all()

    # ORM tree for sidebar rendering
    orm_tree: Dict[str, Any] = {}
    for f in all_folders:
        gid = f.device_group_id or "unknown"
        if gid == "global" or gid.endswith(":global"):
            continue
        dev_name = device_names.get(gid, f"Device {gid[:8]}")
        if gid not in orm_tree:
            orm_tree[gid] = {"name": dev_name, "id": gid, "sections": {"pre": [], "post": [], "default": []}}
        f.rules_count = 0
        sec = f.section.lower() if f.section and f.section.lower() in ('pre', 'post', 'default') else 'pre'
        orm_tree[gid]["sections"][sec].append(f)

    # JSON-serialisable version for JS
    js_tree: Dict[str, Any] = {}
    for gid, info in orm_tree.items():
        js_tree[gid] = {
            "name": info["name"],
            "sections": {
                sec: [{"id": f.id, "name": f.name} for f in flist]
                for sec, flist in info["sections"].items()
            }
        }

    devices_list = [{"device_id": gid, "name": info["name"]} for gid, info in orm_tree.items()]

    return templates.TemplateResponse(request, "analyzer.html", base_ctx(request,
        tree=orm_tree, js_tree=js_tree, devices=devices_list, user=user,
        selected_folder_id=None, current_device_id=None,
        selected_su_id=current_su_id,
    ))


@router.post("/api/v1/analyzer/run")
async def run_analyzer(data: AnalyzerRequest, db: AsyncSession = Depends(get_db)):
    # Collect folder IDs belonging to this device
    folders_stmt = select(Folder).where(Folder.device_group_id == data.device_group_id)
    fr = await db.execute(folders_stmt)
    folders_list = fr.scalars().all()
    folder_names: Dict[str, str] = {f.id: f.name for f in folders_list}
    folder_section: Dict[str, str] = {f.id: (f.section or 'pre').lower() for f in folders_list}

    target_folder_ids = [f.id for f in folders_list]
    if data.folder_id:
        if data.folder_id not in folder_names:
            return JSONResponse({"error": "Folder not found"}, status_code=404)
        target_folder_ids = [data.folder_id]

    if not target_folder_ids:
        return JSONResponse(run_analysis([]))

    stmt = (
        select(CachedRule)
        .where(CachedRule.folder_id.in_(target_folder_ids))
        .order_by(CachedRule.folder_sort_order)
    )
    rows = await db.execute(stmt)
    rules = rows.scalars().all()

    rules_meta = []
    for r in rules:
        rules_meta.append({
            "id":              r.id,
            "ext_id":          r.ext_id,
            "name":            r.name,
            "folder_id":       r.folder_id,
            "folder_name":     folder_names.get(r.folder_id, ""),
            "device_group_id": data.device_group_id,
            "data":            r.data or {},
        })

    result = run_analysis(rules_meta)
    return JSONResponse(result)


@router.get("/api/v1/analyzer/cached")
async def get_cached_analysis(request: Request, db: AsyncSession = Depends(get_db)):
    """Return the latest auto-analysis result stored after the last Sync, scoped to current СУ."""
    user = get_current_user(request)
    current_su_id = user.get('su_id') if user else None
    stmt = select(CachedAnalysis).order_by(CachedAnalysis.analyzed_at.desc())
    if current_su_id:
        stmt = stmt.where(CachedAnalysis.su_id == current_su_id)
    row = (await db.execute(stmt.limit(1))).scalars().first()

    if not row:
        return JSONResponse({"status": "no_data"})

    return JSONResponse({
        "status": "ok",
        "analyzed_at": row.analyzed_at.strftime("%Y-%m-%d %H:%M UTC"),
        "total_rules": row.total_rules,
        "total_issues": row.total_issues,
        **row.result,
    })


# ---------------------------------------------------------------------------
# Change Log
# ---------------------------------------------------------------------------

@router.get("/changelog", response_class=HTMLResponse)
async def changelog_page(request: Request, db: AsyncSession = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login")
    spa = _try_spa()
    if spa: return spa
    return templates.TemplateResponse(request, "changelog.html", base_ctx(request,
        user=user, selected_folder_id=None, current_device_id=None, tree={},
    ))


class ChangelogQueryRequest(BaseModel):
    limit:           int = 100
    offset:          int = 0
    username:        Optional[str] = None
    entity_type:     Optional[str] = None
    action:          Optional[str] = None
    device_group_id: Optional[str] = None
    search:          Optional[str] = None


@router.post("/api/v1/changelog/query")
async def query_changelog(data: ChangelogQueryRequest, db: AsyncSession = Depends(get_db)):
    stmt = select(ChangeLog).order_by(ChangeLog.ts.desc())
    if data.username:
        stmt = stmt.where(ChangeLog.username == data.username)
    if data.entity_type:
        stmt = stmt.where(ChangeLog.entity_type == data.entity_type)
    if data.action:
        stmt = stmt.where(ChangeLog.action == data.action)
    if data.device_group_id:
        stmt = stmt.where(ChangeLog.device_group_id == data.device_group_id)
    if data.search:
        like = f"%{data.search}%"
        stmt = stmt.where(
            ChangeLog.entity_name.ilike(like) | ChangeLog.detail.ilike(like)
        )

    total_stmt = select(func.count()).select_from(stmt.subquery())
    total = (await db.execute(total_stmt)).scalar() or 0

    stmt = stmt.offset(data.offset).limit(data.limit)
    rows = (await db.execute(stmt)).scalars().all()

    items = []
    for r in rows:
        items.append({
            "id":             r.id,
            "ts":             r.ts.isoformat() if r.ts else None,
            "username":       r.username,
            "device_group_id": r.device_group_id,
            "entity_type":    r.entity_type,
            "entity_id":      r.entity_id,
            "entity_name":    r.entity_name,
            "action":         r.action,
            "detail":         r.detail,
        })
    return JSONResponse({"status": "ok", "total": total, "items": items})


# ---------------------------------------------------------------------------
# Policy Diff
# ---------------------------------------------------------------------------

class DeviceDiffRequest(BaseModel):
    device_a: str
    device_b: str


def _rule_signature(data: Dict) -> Dict:
    """Extract comparable fields from rule data."""
    def _ids(field):
        if not field:
            return None
        kind = field.get("kind", "")
        if "ANY" in kind:
            return None
        objs = field.get("objects", [])
        if isinstance(objs, dict):
            objs = objs.get("array", [])
        ids = set()
        for o in objs:
            if isinstance(o, dict):
                ids.add(o.get("id", ""))
        return frozenset(ids) or None

    action = data.get("action", "")
    action = action.split("_")[-1].upper() if "_" in action else action.upper()

    return {
        "action":   action,
        "enabled":  data.get("enabled", True),
        "srcZone":  _ids(data.get("sourceZone")),
        "dstZone":  _ids(data.get("destinationZone")),
        "srcAddr":  _ids(data.get("sourceAddr")),
        "dstAddr":  _ids(data.get("destinationAddr")),
        "service":  _ids(data.get("service")),
    }


def _sig_diff(a: Dict, b: Dict) -> List[str]:
    """Return list of field names that differ between two signatures."""
    diffs = []
    for k in a:
        if a[k] != b[k]:
            diffs.append(k)
    return diffs


@router.get("/diff", response_class=HTMLResponse)
async def diff_page(request: Request, db: AsyncSession = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login")
    spa = _try_spa()
    if spa: return spa

    meta_res = await db.execute(select(DeviceMeta))
    devices = [{"device_id": d.device_id, "name": d.name or d.device_id}
               for d in meta_res.scalars().all() if d.device_id != "global"]

    return templates.TemplateResponse(request, "diff.html", base_ctx(request,
        user=user, devices=devices, selected_folder_id=None, current_device_id=None, tree={},
    ))


@router.post("/api/v1/diff/devices")
async def diff_devices(data: DeviceDiffRequest, db: AsyncSession = Depends(get_db)):
    """Compare cached rules between two device groups."""

    async def _load(device_id: str):
        folders_stmt = select(Folder).where(Folder.device_group_id == device_id)
        folder_rows = (await db.execute(folders_stmt)).scalars().all()
        folder_map = {f.id: f.name for f in folder_rows}

        rules_stmt = select(CachedRule).where(
            CachedRule.folder_id.in_(list(folder_map.keys()))
        ).order_by(CachedRule.folder_sort_order)
        rules = (await db.execute(rules_stmt)).scalars().all()
        return {
            r.name: {
                "id":     r.id,
                "ext_id": r.ext_id,
                "name":   r.name,
                "folder": folder_map.get(r.folder_id, ""),
                "sig":    _rule_signature(r.data or {}),
                "data":   r.data or {},
            }
            for r in rules
        }

    rules_a = await _load(data.device_a)
    rules_b = await _load(data.device_b)

    names_a = set(rules_a)
    names_b = set(rules_b)

    only_a = []
    for name in sorted(names_a - names_b):
        r = rules_a[name]
        only_a.append({"name": name, "folder": r["folder"],
                        "action": r["sig"]["action"], "enabled": r["sig"]["enabled"]})

    only_b = []
    for name in sorted(names_b - names_a):
        r = rules_b[name]
        only_b.append({"name": name, "folder": r["folder"],
                        "action": r["sig"]["action"], "enabled": r["sig"]["enabled"]})

    # Build object name lookup for both devices
    all_obj_ids = set()
    for r in {**rules_a, **rules_b}.values():
        for field in r["sig"].values():
            if isinstance(field, frozenset):
                all_obj_ids.update(field)
    obj_name_map: dict = {}
    if all_obj_ids:
        obj_rows = (await db.execute(
            select(CachedObject.ext_id, CachedObject.name)
            .where(CachedObject.ext_id.in_(list(all_obj_ids)))
        )).all()
        obj_name_map = {row.ext_id: row.name for row in obj_rows}

    def _fmt(v):
        if v is None:
            return "ANY"
        if isinstance(v, bool):
            return "enabled" if v else "disabled"
        if isinstance(v, frozenset):
            names = sorted(obj_name_map.get(i, i[:8]) for i in v)
            return ", ".join(names) if names else "ANY"
        return str(v)

    changed = []
    for name in sorted(names_a & names_b):
        ra, rb = rules_a[name], rules_b[name]
        diffs = _sig_diff(ra["sig"], rb["sig"])
        if diffs:
            changed.append({
                "name":     name,
                "folder_a": ra["folder"],
                "folder_b": rb["folder"],
                "diffs":    diffs,
                "a": {k: _fmt(v) for k, v in ra["sig"].items()},
                "b": {k: _fmt(v) for k, v in rb["sig"].items()},
            })

    return JSONResponse({
        "status":    "ok",
        "total_a":   len(rules_a),
        "total_b":   len(rules_b),
        "only_a":    only_a,
        "only_b":    only_b,
        "changed":   changed,
        "identical": len(names_a & names_b) - len(changed),
    })


@router.get("/api/v1/diff/modified")
async def diff_modified(device_group_id: str = Query(...), db: AsyncSession = Depends(get_db)):
    """Return rules marked as modified (pending deploy)."""
    folders_stmt = select(Folder).where(Folder.device_group_id == device_group_id)
    folder_map = {f.id: f.name for f in (await db.execute(folders_stmt)).scalars().all()}

    if not folder_map:
        return JSONResponse({"status": "ok", "rules": []})

    stmt = select(CachedRule).where(
        CachedRule.folder_id.in_(list(folder_map.keys())),
        CachedRule.is_modified == True,
    )
    rules = (await db.execute(stmt)).scalars().all()

    return JSONResponse({
        "status": "ok",
        "rules": [
            {
                "id":     r.id,
                "ext_id": r.ext_id,
                "name":   r.name,
                "folder": folder_map.get(r.folder_id, ""),
                "modified_at": r.modified_at,
            }
            for r in rules
        ],
    })


# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------

@router.get("/dashboard", response_class=HTMLResponse)
async def dashboard_page(
    request: Request,
    device_id: str = Query(None),
    su_id: str = Query(None),
    db: AsyncSession = Depends(get_db),
):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login")
    spa = _try_spa()
    if spa: return spa

    import app.state as _state
    # Derive su_id from scoped device_id prefix when not explicitly provided
    if not su_id and device_id and ":" in device_id:
        su_id = device_id.split(":")[0]
    # Resolve which СУ to display: explicit ?su_id param → session su_id → first available
    current_su_id = su_id or user.get('su_id') or (
        _state.su_list[0]["id"] if _state.su_list else None
    )

    # All devices (for sidebar list) — scoped to current СУ
    meta_stmt = select(DeviceMeta)
    if current_su_id:
        meta_stmt = meta_stmt.where(DeviceMeta.su_id == current_su_id)
    all_meta = [m for m in (await db.execute(meta_stmt)).scalars().all()
                if not m.device_id.endswith(":global") and m.device_id != "global"]
    device_names = {m.device_id: m.name for m in all_meta}

    # --- Build folder → device map (needed for rule stats) — scoped to current СУ ---
    folders_stmt = select(Folder).order_by(Folder.device_group_id)
    if current_su_id:
        folders_stmt = folders_stmt.where(Folder.su_id == current_su_id)
    folder_to_device: Dict[str, str] = {}
    for f in (await db.execute(folders_stmt)).scalars().all():
        if f.device_group_id and not f.device_group_id.endswith(":global") and f.device_group_id != "global":
            folder_to_device[f.id] = f.device_group_id

    # --- Helper: build stats for one or all devices (always scoped to current_su_id) ---
    async def _build_stats(gid_filter: Optional[str]) -> Dict:
        stats = {"sec_total": 0, "sec_modified": 0, "nat_total": 0, "nat_modified": 0, "objects": 0}

        rule_q = select(CachedRule)
        for r in (await db.execute(rule_q)).scalars().all():
            gid = folder_to_device.get(r.folder_id or "")
            if not gid:  # folder not in current su_id scope
                continue
            if gid_filter and gid != gid_filter:
                continue
            stats["sec_total"] += 1
            if r.is_modified:
                stats["sec_modified"] += 1

        nat_q = select(CachedNatRule)
        if gid_filter:
            nat_q = nat_q.where(CachedNatRule.device_group_id == gid_filter)
        elif current_su_id:
            nat_q = nat_q.where(CachedNatRule.su_id == current_su_id)
        for r in (await db.execute(nat_q)).scalars().all():
            stats["nat_total"] += 1
            if r.is_modified:
                stats["nat_modified"] += 1

        obj_q = select(CachedObject)
        if gid_filter:
            obj_q = obj_q.where(CachedObject.device_group_id == gid_filter)
        elif current_su_id:
            obj_q = obj_q.where(CachedObject.su_id == current_su_id)
        stats["objects"] = len((await db.execute(obj_q)).scalars().all())
        return stats

    # --- Changelog (scoped to current su_id via device_group_id prefix) ---
    async def _build_changelog(gid_filter: Optional[str], limit: int = 20):
        q = select(ChangeLog).order_by(ChangeLog.ts.desc())
        if gid_filter:
            q = q.where(ChangeLog.device_group_id == gid_filter)
        elif current_su_id:
            q = q.where(ChangeLog.device_group_id.like(f"{current_su_id}:%"))
        q = q.limit(limit)
        result = []
        for c in (await db.execute(q)).scalars().all():
            gid = c.device_group_id or ""
            result.append({
                "ts":          c.ts.strftime("%Y-%m-%d %H:%M") if c.ts else "",
                "username":    c.username,
                "entity_type": c.entity_type,
                "entity_name": c.entity_name or "",
                "action":      c.action,
                "device_name": device_names.get(gid, gid[:8] if gid else "—"),
            })
        return result

    # --- Analyzer: filter cached result for device (scoped to current su_id) ---
    async def _build_analysis(gid_filter: Optional[str]) -> Optional[Dict]:
        try:
            stmt = select(CachedAnalysis).order_by(CachedAnalysis.analyzed_at.desc())
            if current_su_id:
                stmt = stmt.where(CachedAnalysis.su_id == current_su_id)
            row = (await db.execute(stmt.limit(1))).scalars().first()
            if not row:
                return None
            result = row.result or {}
            if not gid_filter:
                return {"total_issues": row.total_issues, "analyzed_at": row.analyzed_at.strftime("%H:%M")}
            def _filt(lst): return [x for x in (lst or []) if x.get("device") == gid_filter]
            dis = _filt(result.get("disabled", []))
            broad = _filt(result.get("too_broad", []))
            shadow = _filt(result.get("shadowed", []))
            redund = _filt(result.get("redundant", []))
            return {
                "total_issues": len(dis) + len(broad) + len(shadow) + len(redund),
                "disabled": dis, "too_broad": broad, "shadowed": shadow, "redundant": redund,
                "analyzed_at": row.analyzed_at.strftime("%H:%M"),
            }
        except Exception as e:
            logger.warning(f"[dashboard] _build_analysis failed: {e}")
            return None

    # ── Global view (no device selected) ──
    if not device_id:
        try:
            changelog = await _build_changelog(None, limit=20)
            chlog_q = select(func.count()).select_from(ChangeLog)
            if current_su_id:
                chlog_q = chlog_q.where(ChangeLog.device_group_id.like(f"{current_su_id}:%"))
            total_chlog = (await db.execute(chlog_q)).scalar() or 0
        except Exception as e:
            logger.warning(f"[dashboard] changelog query failed: {e}")
            changelog, total_chlog = [], 0
        analysis = await _build_analysis(None)

        return templates.TemplateResponse(request, "dashboard.html", base_ctx(request,
            user=user, devices=all_meta, selected_device_id=None,
            view="global",
            device_count=len(all_meta),
            changelog=changelog, total_chlog=total_chlog,
            analysis=analysis,
            selected_su_id=current_su_id,
        ))

    # ── Device view ──
    dev_name = device_names.get(device_id, device_id[:8])
    stats = await _build_stats(device_id)
    changelog = await _build_changelog(device_id, limit=20)
    analysis = await _build_analysis(device_id)

    # Modified rules for this device
    modified_rules = []
    for r in (await db.execute(select(CachedRule))).scalars().all():
        if r.is_modified and folder_to_device.get(r.folder_id or "") == device_id:
            modified_rules.append({
                "name": r.name,
                "modified_at": r.modified_at or "",
                "ext_id": r.ext_id or "",
                "device_group_id": device_id,
            })

    return templates.TemplateResponse(request, "dashboard.html", base_ctx(request,
        user=user, devices=all_meta, selected_device_id=device_id,
        view="device",
        device_name=dev_name, device_id=device_id,
        stats=stats, changelog=changelog, analysis=analysis,
        modified_rules=modified_rules,
        selected_su_id=current_su_id,
    ))


# ---------------------------------------------------------------------------
# Smart Rule Search — Phase 7.10 + Object Usage Map — Phase 7.5
# ---------------------------------------------------------------------------

class SearchRequest(BaseModel):
    query: str
    mode: str = "any"      # any | name | ip | port | object
    device_group_id: str = ""


def _rule_action_norm(data: Dict) -> str:
    raw = data.get("action", "")
    return raw.split("_")[-1].upper() if "_" in raw else raw.upper()


def _extract_obj_ids_from_rule_data(data: Dict) -> set:
    """Collect all object UUIDs referenced in a rule's field values."""
    ids = set()
    for field_key in ("sourceAddr", "destinationAddr", "service",
                      "sourceZone", "destinationZone"):
        field = data.get(field_key)
        if not field:
            continue
        objects = field.get("objects", [])
        if isinstance(objects, dict):
            objects = objects.get("array", [])
        for item in objects:
            if isinstance(item, str):
                ids.add(item)
            elif isinstance(item, dict):
                if "id" in item:
                    ids.add(item["id"])
                else:
                    for v in item.values():
                        if isinstance(v, dict) and "id" in v:
                            ids.add(v["id"])
    return ids


async def _do_search(
    db: AsyncSession,
    query: str,
    mode: str,
    device_group_id: str,
    device_names: Dict[str, str],
    folder_map: Dict[str, Any],   # folder_id → {name, device_group_id}
) -> List[Dict]:
    """Core search logic. Returns list of result dicts."""
    q = query.strip()
    if not q:
        return []

    results: List[Dict] = []
    seen_ids: set = set()

    def _folder_info(folder_id):
        fi = folder_map.get(folder_id or "", {})
        return fi.get("name", ""), fi.get("device_group_id", "")

    def _add_rule(rule, match_reason):
        if rule.id in seen_ids:
            return
        seen_ids.add(rule.id)
        folder_name, gid = _folder_info(rule.folder_id)
        d = rule.data or {}
        results.append({
            "type":        "rule",
            "id":          rule.id,
            "ext_id":      rule.ext_id,
            "name":        rule.name,
            "enabled":     d.get("enabled", True),
            "action":      _rule_action_norm(d),
            "folder":      folder_name,
            "device":      device_names.get(gid, gid[:12] if gid else ""),
            "device_id":   gid,
            "match":       match_reason,
            "folder_id":   rule.folder_id or "",
        })

    # Base rule query (filter by device if specified)
    base_rule_stmt = select(CachedRule)
    if device_group_id:
        # filter by folders of that device
        dev_folder_ids = [fid for fid, fi in folder_map.items()
                         if fi.get("device_group_id") == device_group_id]
        if dev_folder_ids:
            base_rule_stmt = base_rule_stmt.where(CachedRule.folder_id.in_(dev_folder_ids))
        else:
            return []

    # --- NAME search ---
    if mode in ("any", "name"):
        stmt = base_rule_stmt.where(CachedRule.name.ilike(f"%{q}%"))
        for r in (await db.execute(stmt)).scalars().all():
            _add_rule(r, f"Name matches «{q}»")

    # --- OBJECT ext_id search (Object Usage Map) ---
    if mode == "object":
        all_rules = (await db.execute(base_rule_stmt)).scalars().all()
        for r in all_rules:
            ids = _extract_obj_ids_from_rule_data(r.data or {})
            if q in ids:
                _add_rule(r, f"References object {q[:16]}…")
        return results[:200]

    # --- IP / PORT search — first resolve matching objects ---
    matching_obj_ids: set = set()

    if mode in ("any", "ip"):
        obj_stmt = select(CachedObject).where(
            CachedObject.category == "net",
            cast(CachedObject.data, Text).ilike(f"%{q}%"),
        )
        if device_group_id:
            obj_stmt = obj_stmt.where(
                or_(CachedObject.device_group_id == device_group_id,
                    CachedObject.device_group_id == "global")
            )
        for o in (await db.execute(obj_stmt)).scalars().all():
            matching_obj_ids.add(o.ext_id)

    if mode in ("any", "port"):
        try:
            port_num = int(q)
            obj_stmt = select(CachedObject).where(
                CachedObject.category == "service",
                cast(CachedObject.data, Text).ilike(f"%{port_num}%"),
            )
            if device_group_id:
                obj_stmt = obj_stmt.where(
                    or_(CachedObject.device_group_id == device_group_id,
                        CachedObject.device_group_id == "global")
                )
            for o in (await db.execute(obj_stmt)).scalars().all():
                matching_obj_ids.add(o.ext_id)
        except ValueError:
            pass  # not a number, skip port search

    # Now find rules referencing those objects
    if matching_obj_ids:
        all_rules = (await db.execute(base_rule_stmt)).scalars().all()
        for r in all_rules:
            ids = _extract_obj_ids_from_rule_data(r.data or {})
            common = ids & matching_obj_ids
            if common:
                _add_rule(r, f"References matching object(s)")

    return results[:300]


@router.get("/search", response_class=HTMLResponse)
async def search_page(
    request: Request,
    q: str = Query(""),
    mode: str = Query("any"),
    device_id: str = Query(""),
    db: AsyncSession = Depends(get_db),
):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login")
    spa = _try_spa()
    if spa: return spa

    meta_res = await db.execute(select(DeviceMeta))
    device_names = {o.device_id: o.name for o in meta_res.scalars().all()}

    folders_res = await db.execute(select(Folder).order_by(Folder.device_group_id, Folder.sort_order))
    all_folders = folders_res.scalars().all()

    orm_tree: Dict[str, Any] = {}
    folder_map: Dict[str, Any] = {}
    for f in all_folders:
        gid = f.device_group_id or "unknown"
        if gid == "global":
            continue
        folder_map[f.id] = {"name": f.name, "device_group_id": gid}
        dev_name = device_names.get(gid, f"Device {gid[:8]}")
        if gid not in orm_tree:
            orm_tree[gid] = {"name": dev_name, "id": gid, "sections": {"pre": [], "post": [], "default": []}}
        f.rules_count = 0
        sec = f.section.lower() if f.section and f.section.lower() in ('pre', 'post', 'default') else 'pre'
        orm_tree[gid]["sections"][sec].append(f)

    results = []
    if q:
        results = await _do_search(db, q, mode, device_id, device_names, folder_map)

    devices_list = [{"device_id": gid, "name": info["name"]} for gid, info in orm_tree.items()]

    return templates.TemplateResponse(request, "search.html", base_ctx(request,
        tree=orm_tree, user=user,
        selected_folder_id=None, current_device_id=None,
        q=q, mode=mode, device_id=device_id, devices=devices_list, results=results,
    ))


@router.post("/api/v1/search")
async def search_api(data: SearchRequest, db: AsyncSession = Depends(get_db)):
    meta_res = await db.execute(select(DeviceMeta))
    device_names = {o.device_id: o.name for o in meta_res.scalars().all()}

    folders_res = await db.execute(select(Folder))
    folder_map = {
        f.id: {"name": f.name, "device_group_id": f.device_group_id or ""}
        for f in folders_res.scalars().all()
        if (f.device_group_id or "") != "global"
    }

    results = await _do_search(
        db, data.query, data.mode, data.device_group_id,
        device_names, folder_map
    )
    return JSONResponse({"status": "ok", "results": results, "count": len(results)})


@router.get("/api/v1/object/usage")
async def object_usage(ext_id: str = Query(...), db: AsyncSession = Depends(get_db)):
    """Return all rules that reference the given object ext_id."""
    meta_res = await db.execute(select(DeviceMeta))
    device_names = {o.device_id: o.name for o in meta_res.scalars().all()}

    folders_res = await db.execute(select(Folder))
    folder_map = {
        f.id: {"name": f.name, "device_group_id": f.device_group_id or ""}
        for f in folders_res.scalars().all()
        if (f.device_group_id or "") != "global"
    }

    results = await _do_search(db, ext_id, "object", "", device_names, folder_map)
    return JSONResponse({"status": "ok", "results": results, "count": len(results)})


# ---------------------------------------------------------------------------
# Export — Phase 7.8
# ---------------------------------------------------------------------------

def _rule_action_color(action: str) -> str:
    """Return hex color for action badge in Excel."""
    a = action.upper()
    if a in ("ALLOW", "PASS"):
        return "FF22C55E"  # green
    if a in ("DENY", "DROP"):
        return "FFEF4444"  # red
    return "FF94A3B8"      # gray


@router.get("/api/v1/export/rules/xlsx")
async def export_rules_xlsx(
    device_group_id: str = Query(...),
    request: Request = None,
    db: AsyncSession = Depends(get_db),
):
    """Export all Security Rules for a device to a formatted Excel file."""
    # Load folder + rule data
    folders_stmt = select(Folder).where(Folder.device_group_id == device_group_id).order_by(Folder.sort_order)
    folders = (await db.execute(folders_stmt)).scalars().all()
    folder_map = {f.id: f for f in folders}

    if not folders:
        return JSONResponse({"error": "No folders for this device"}, status_code=404)

    folder_ids = [f.id for f in folders]
    rules_stmt = select(CachedRule).where(CachedRule.folder_id.in_(folder_ids)).order_by(
        CachedRule.folder_id, CachedRule.folder_sort_order
    )
    rules = (await db.execute(rules_stmt)).scalars().all()

    # Object cache for name resolution
    obj_stmt = select(CachedObject).where(
        or_(CachedObject.device_group_id == device_group_id, CachedObject.device_group_id == "global")
    )
    obj_map = {o.ext_id: o.name for o in (await db.execute(obj_stmt)).scalars().all()}

    def _resolve_ids(field: Dict) -> str:
        if not field:
            return "Any"
        kind = field.get("kind", "")
        if "ANY" in kind:
            return "Any"
        objects = field.get("objects", [])
        if isinstance(objects, dict):
            objects = objects.get("array", [])
        names = []
        for item in objects:
            if isinstance(item, str):
                names.append(obj_map.get(item, item[:8]))
            elif isinstance(item, dict):
                oid = item.get("id") or next(
                    (v.get("id") for v in item.values() if isinstance(v, dict)), None
                )
                names.append(obj_map.get(oid, oid[:8] if oid else "?") if oid else "?")
        return ", ".join(names) if names else "Any"

    # Build workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Security Rules"

    # Styles
    hdr_font   = Font(name="Calibri", bold=True, color="FFFFFFFF", size=10)
    hdr_fill   = PatternFill("solid", fgColor="FF0F172A")
    wrap_align = Alignment(wrap_text=True, vertical="top")
    ctr_align  = Alignment(horizontal="center", vertical="top")

    headers = ["#", "Folder", "Rule Name", "Enabled", "Action",
               "Src Zone", "Source", "Dst Zone", "Destination", "Service", "Comment"]
    col_widths = [5, 18, 30, 8, 8, 14, 30, 14, 30, 20, 30]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = ctr_align
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "A2"

    row_idx = 2
    for i, rule in enumerate(rules, 1):
        d = rule.data or {}
        action_raw = d.get("action", "")
        action = action_raw.split("_")[-1].upper() if "_" in action_raw else action_raw.upper()
        enabled = d.get("enabled", True)

        folder_name = folder_map.get(rule.folder_id, None)
        folder_str  = folder_name.name if folder_name else ""

        row_data = [
            i,
            folder_str,
            rule.name or "",
            "Yes" if enabled else "No",
            action,
            _resolve_ids(d.get("sourceZone")),
            _resolve_ids(d.get("sourceAddr")),
            _resolve_ids(d.get("destinationZone")),
            _resolve_ids(d.get("destinationAddr")),
            _resolve_ids(d.get("service")),
            d.get("description", "") or "",
        ]

        # Alternating row background
        row_fill = PatternFill("solid", fgColor="FF1E293B" if i % 2 == 0 else "FF0F172A")

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = row_fill
            cell.alignment = wrap_align
            cell.font = Font(name="Calibri", size=10, color="FFE2E8F0")

        # Color the Action cell
        action_cell = ws.cell(row=row_idx, column=5)
        action_cell.fill = PatternFill("solid", fgColor=_rule_action_color(action))
        action_cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFFFF")
        action_cell.alignment = ctr_align

        ws.row_dimensions[row_idx].height = 30
        row_idx += 1

    # Add auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Serialize to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    meta_res = await db.execute(select(DeviceMeta).where(DeviceMeta.device_id == device_group_id))
    dev = meta_res.scalar_one_or_none()
    dev_name = (dev.name if dev else device_group_id[:12]).replace(" ", "_")
    filename = f"rules_{dev_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/api/v1/export/folders/yaml")
async def export_folders_yaml(
    device_group_id: str = Query(...),
    db: AsyncSession = Depends(get_db),
):
    """Export virtual folder structure (including rule ordering) to YAML."""
    folders_stmt = select(Folder).where(Folder.device_group_id == device_group_id).order_by(Folder.sort_order)
    folders = (await db.execute(folders_stmt)).scalars().all()

    folder_ids = [f.id for f in folders]
    rules_stmt = select(CachedRule).where(CachedRule.folder_id.in_(folder_ids)).order_by(
        CachedRule.folder_id, CachedRule.folder_sort_order
    )
    rules = (await db.execute(rules_stmt)).scalars().all()

    # Group rules by folder
    rules_by_folder: Dict[str, List] = {}
    for r in rules:
        rules_by_folder.setdefault(r.folder_id, []).append({
            "id":    r.id,
            "ext_id": r.ext_id,
            "name":  r.name,
            "sort_order": r.folder_sort_order,
        })

    meta_res = await db.execute(select(DeviceMeta).where(DeviceMeta.device_id == device_group_id))
    dev = meta_res.scalar_one_or_none()

    structure = {
        "device_group_id": device_group_id,
        "device_name":     dev.name if dev else "",
        "exported_at":     datetime.now(timezone.utc).isoformat(),
        "sections": {
            "pre":     [],
            "default": [],
            "post":    [],
        }
    }

    for f in folders:
        section = (f.section or "default").lower()
        if section not in structure["sections"]:
            section = "default"
        structure["sections"][section].append({
            "id":     f.id,
            "name":   f.name,
            "sort_order": f.sort_order,
            "rules":  rules_by_folder.get(f.id, []),
        })

    yaml_str = yaml.dump(structure, allow_unicode=True, default_flow_style=False, sort_keys=False)

    dev_name = (dev.name if dev else device_group_id[:12]).replace(" ", "_")
    filename = f"folders_{dev_name}_{datetime.now().strftime('%Y%m%d')}.yaml"

    return StreamingResponse(
        io.BytesIO(yaml_str.encode("utf-8")),
        media_type="text/yaml",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/api/v1/import/folders/yaml")
async def import_folders_yaml(
    request: Request,
    device_group_id: str = Form(...),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    """Restore virtual folder structure from a previously exported YAML file."""
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error", "message": "Unauthorized"}, status_code=401)

    content = await file.read()
    try:
        data = yaml.safe_load(content.decode("utf-8"))
    except Exception as e:
        return JSONResponse({"status": "error", "message": f"YAML parse error: {e}"}, status_code=400)

    sections = data.get("sections") if isinstance(data, dict) else None
    if not sections:
        return JSONResponse({"status": "error", "message": "Invalid YAML: 'sections' key missing"}, status_code=400)

    # Load existing folders keyed by (section, name)
    folders_res = await db.execute(select(Folder).where(Folder.device_group_id == device_group_id))
    existing_folders: Dict[tuple, Folder] = {(f.section, f.name): f for f in folders_res.scalars().all()}

    # Load all rules for this device keyed by ext_id
    fids = list({f.id for f in existing_folders.values()})
    all_rules: Dict[str, CachedRule] = {}
    if fids:
        rules_res = await db.execute(select(CachedRule).where(CachedRule.folder_id.in_(fids)))
        all_rules = {r.ext_id: r for r in rules_res.scalars().all()}

    folders_created = 0
    folders_updated = 0
    rules_assigned  = 0

    for section_name, folder_list in sections.items():
        if not isinstance(folder_list, list):
            continue
        for folder_data in folder_list:
            fname = str(folder_data.get("name", "")).strip()
            if not fname:
                continue
            fsort = int(folder_data.get("sort_order") or 0)

            key = (section_name, fname)
            if key in existing_folders:
                folder = existing_folders[key]
                folder.sort_order = fsort
                folders_updated += 1
            else:
                folder = Folder(
                    id=str(uuid.uuid4()),
                    name=fname,
                    device_group_id=device_group_id,
                    section=section_name,
                    sort_order=fsort,
                )
                db.add(folder)
                existing_folders[key] = folder
                folders_created += 1

            for rule_data in (folder_data.get("rules") or []):
                ext_id = rule_data.get("ext_id")
                if not ext_id or ext_id not in all_rules:
                    continue
                rule = all_rules[ext_id]
                rule.folder_id = folder.id
                rule.folder_sort_order = int(rule_data.get("sort_order") or 0)
                rules_assigned += 1

    await db.commit()
    return JSONResponse({
        "status": "ok",
        "folders_created": folders_created,
        "folders_updated": folders_updated,
        "rules_assigned":  rules_assigned,
    })


@router.get("/api/v1/export/rules/html")
async def export_rules_html(
    request: Request,
    device_group_id: str = Query(...),
    db: AsyncSession = Depends(get_db),
):
    """Generate a printable HTML policy report (browser Print → Save as PDF)."""
    if not get_current_user(request):
        return HTMLResponse("<h1>Unauthorized</h1>", status_code=401)

    meta_res = await db.execute(select(DeviceMeta).where(DeviceMeta.device_id == device_group_id))
    dev = meta_res.scalar_one_or_none()
    dev_name = dev.name if dev else device_group_id

    folders_res = await db.execute(
        select(Folder).where(Folder.device_group_id == device_group_id).order_by(Folder.sort_order)
    )
    folders = folders_res.scalars().all()
    fid_map = {f.id: f for f in folders}

    rules_res = await db.execute(
        select(CachedRule)
        .where(CachedRule.folder_id.in_(list(fid_map.keys())))
        .order_by(CachedRule.folder_id, CachedRule.folder_sort_order)
    )
    rules = rules_res.scalars().all()

    name_map: Dict[str, str] = {k: v for k, v in GLOBAL_NAME_MAP.items()}
    if not name_map:
        nm_res = await db.execute(select(CachedObject.ext_id, CachedObject.name))
        for row in nm_res:
            name_map[row[0]] = row[1]

    def _resolve(field: Optional[Dict]) -> str:
        if not field:
            return "ANY"
        kind = str(field.get("kind", ""))
        if "ANY" in kind:
            return "ANY"
        objects = field.get("objects", [])
        if isinstance(objects, dict):
            ids = objects.get("array", [])
        elif isinstance(objects, list):
            ids = []
            for o in objects:
                if isinstance(o, dict):
                    ids.append(o.get("id", ""))
                elif isinstance(o, str):
                    ids.append(o)
        else:
            ids = []
        names = [name_map.get(i, i[:8] + "…") for i in ids if i]
        return ", ".join(names) if names else "ANY"

    ACTION_COLOR = {
        "SECURITY_RULE_ACTION_ALLOW": "#16a34a",
        "SECURITY_RULE_ACTION_DENY":  "#dc2626",
        "SECURITY_RULE_ACTION_DROP":  "#d97706",
    }

    rows_by_folder: Dict[str, list] = {}
    for r in rules:
        rows_by_folder.setdefault(r.folder_id, []).append(r)

    # Build HTML
    sections_html = []
    for folder in folders:
        folder_rules = rows_by_folder.get(folder.id, [])
        if not folder_rules:
            continue
        row_rows = []
        for i, r in enumerate(folder_rules, 1):
            d = r.data or {}
            action = d.get("action", "")
            color  = ACTION_COLOR.get(action, "#6b7280")
            action_lbl = action.replace("SECURITY_RULE_ACTION_", "")
            enabled_lbl = "" if d.get("enabled", True) else " ⛔"
            row_rows.append(f"""
            <tr class="{'disabled-row' if not d.get('enabled', True) else ''}">
                <td>{i}</td>
                <td><b>{r.name}{enabled_lbl}</b></td>
                <td style="color:{color};font-weight:700">{action_lbl}</td>
                <td>{_resolve(d.get('sourceZone'))}</td>
                <td>{_resolve(d.get('destinationZone'))}</td>
                <td>{_resolve(d.get('sourceAddr'))}</td>
                <td>{_resolve(d.get('destinationAddr'))}</td>
                <td>{_resolve(d.get('service'))}</td>
                <td style="color:#6b7280;font-size:11px">{d.get('description','')[:60]}</td>
            </tr>""")

        sections_html.append(f"""
        <div class="folder-block">
            <div class="folder-title">{folder.name}
                <span class="folder-sec">{folder.section or ''}</span>
                <span class="folder-cnt">{len(folder_rules)} rules</span>
            </div>
            <table>
                <thead><tr>
                    <th>#</th><th>Name</th><th>Action</th>
                    <th>Src Zone</th><th>Dst Zone</th>
                    <th>Src Addr</th><th>Dst Addr</th>
                    <th>Service</th><th>Description</th>
                </tr></thead>
                <tbody>{''.join(row_rows)}</tbody>
            </table>
        </div>""")

    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>Policy Report — {dev_name}</title>
<style>
  body {{ font-family: Arial, sans-serif; font-size: 12px; color: #111; margin: 20px; }}
  h1 {{ font-size: 18px; margin-bottom: 4px; }}
  .meta {{ color: #555; font-size: 11px; margin-bottom: 20px; }}
  .folder-block {{ margin-bottom: 24px; page-break-inside: avoid; }}
  .folder-title {{ font-size: 14px; font-weight: 700; background: #f1f5f9; padding: 6px 10px;
                   border-left: 4px solid #3b82f6; margin-bottom: 0; }}
  .folder-sec {{ font-size: 10px; font-weight: 400; color: #64748b; margin-left: 8px; text-transform: uppercase; }}
  .folder-cnt {{ font-size: 10px; color: #94a3b8; float: right; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 11px; }}
  th {{ background: #e2e8f0; padding: 5px 6px; text-align: left; font-size: 10px;
        text-transform: uppercase; letter-spacing: .04em; color: #475569; border: 1px solid #cbd5e1; }}
  td {{ padding: 4px 6px; border: 1px solid #e2e8f0; vertical-align: top; }}
  tr:nth-child(even) td {{ background: #f8fafc; }}
  .disabled-row td {{ opacity: .5; }}
  .print-btn {{ position: fixed; top: 12px; right: 12px; padding: 8px 18px;
                background: #3b82f6; color: #fff; border: none; border-radius: 6px;
                font-size: 13px; font-weight: 700; cursor: pointer; }}
  @media print {{
    .print-btn {{ display: none; }}
    body {{ margin: 0; }}
  }}
</style>
</head>
<body>
<button class="print-btn" onclick="window.print()">🖨 Print / Save PDF</button>
<h1>Security Policy — {dev_name}</h1>
<div class="meta">Generated: {now} &nbsp;|&nbsp; Device group: {device_group_id} &nbsp;|&nbsp; Total folders: {len(folders)}</div>
{''.join(sections_html)}
</body>
</html>"""

    return HTMLResponse(html)


# ---------------------------------------------------------------------------
# Rule Templates — Phase 7.4
# ---------------------------------------------------------------------------

class TemplateSaveRequest(BaseModel):
    rule_id: str          # local CachedRule id to snapshot
    name: str
    description: str = ""


class TemplateApplyRequest(BaseModel):
    template_id: str
    folder_id: str        # target folder (CachedRule folder)


class TemplateDeleteRequest(BaseModel):
    template_ids: List[str]


@router.get("/templates", response_class=HTMLResponse)
async def templates_page(request: Request, db: AsyncSession = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login")
    spa = _try_spa()
    if spa: return spa

    meta_res = await db.execute(select(DeviceMeta))
    device_names = {o.device_id: o.name for o in meta_res.scalars().all()}

    folders_res = await db.execute(select(Folder).order_by(Folder.device_group_id, Folder.sort_order))
    all_folders = folders_res.scalars().all()

    orm_tree: Dict[str, Any] = {}
    for f in all_folders:
        gid = f.device_group_id or "unknown"
        if gid == "global":
            continue
        dev_name = device_names.get(gid, f"Device {gid[:8]}")
        if gid not in orm_tree:
            orm_tree[gid] = {"name": dev_name, "id": gid, "sections": {"pre": [], "post": [], "default": []}}
        f.rules_count = 0
        sec = f.section.lower() if f.section and f.section.lower() in ('pre', 'post', 'default') else 'pre'
        orm_tree[gid]["sections"][sec].append(f)

    tmpl_res = await db.execute(select(RuleTemplate).order_by(RuleTemplate.created_at.desc()))
    templates_list = []
    for t in tmpl_res.scalars().all():
        d = t.data or {}
        action = d.get("action", "")
        action = action.split("_")[-1].upper() if "_" in action else action.upper()
        templates_list.append({
            "id":          t.id,
            "name":        t.name,
            "description": t.description or "",
            "created_by":  t.created_by or "",
            "created_at":  t.created_at.strftime("%Y-%m-%d %H:%M") if t.created_at else "",
            "action":      action,
            "data":        d,
        })

    js_tree: Dict[str, Any] = {}
    for gid, info in orm_tree.items():
        js_tree[gid] = {
            "name": info["name"],
            "sections": {
                sec: [{"id": f.id, "name": f.name} for f in flist]
                for sec, flist in info["sections"].items()
            }
        }

    return templates.TemplateResponse(request, "templates.html", base_ctx(request,
        tree=orm_tree, user=user,
        selected_folder_id=None, current_device_id=None,
        templates=templates_list, js_tree=js_tree,
    ))


@router.post("/api/v1/templates/save")
async def save_template(request: Request, data: TemplateSaveRequest, db: AsyncSession = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error", "message": "Unauthorized"}, status_code=401)

    rule = await db.get(CachedRule, data.rule_id)
    if not rule:
        return JSONResponse({"status": "error", "message": "Rule not found"}, status_code=404)

    # Strip position/id fields from snapshot; keep policy fields
    snap = {k: v for k, v in (rule.data or {}).items()
            if k not in ("id", "position", "globalPosition")}
    snap["name"] = data.name  # template inherits the new name

    tmpl = RuleTemplate(
        id=str(uuid.uuid4()),
        name=data.name,
        description=data.description,
        created_by=user.get("username", ""),
        data=snap,
    )
    db.add(tmpl)
    await db.commit()
    return JSONResponse({"status": "ok", "id": tmpl.id, "name": tmpl.name})


@router.post("/api/v1/templates/apply")
async def apply_template(request: Request, data: TemplateApplyRequest, db: AsyncSession = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error", "message": "Unauthorized"}, status_code=401)

    tmpl = await db.get(RuleTemplate, data.template_id)
    if not tmpl:
        return JSONResponse({"status": "error", "message": "Template not found"}, status_code=404)

    folder = await db.get(Folder, data.folder_id)
    if not folder:
        return JSONResponse({"status": "error", "message": "Folder not found"}, status_code=404)

    snap = dict(tmpl.data or {})
    section = (folder.section or "pre").lower()
    snap["name"] = tmpl.name
    snap["deviceGroupId"] = folder.device_group_id
    snap["precedence"] = section
    snap["position"] = 1
    snap.pop("id", None)

    client = NGFWClient(user['host'], verify_ssl=False)
    try:
        await client.login(user['username'], user['password'])
        res = await client.create_rule(snap)
        ext_id = res.get("id") or res.get("securityRule", {}).get("id")
        if not ext_id:
            raise RuntimeError(f"API did not return rule ID: {res}")

        stmt = select(func.max(CachedRule.folder_sort_order)).where(CachedRule.folder_id == data.folder_id)
        max_pos = (await db.execute(stmt)).scalar() or 0

        db.add(CachedRule(
            id=str(uuid.uuid4()),
            ext_id=ext_id,
            name=tmpl.name,
            folder_id=data.folder_id,
            folder_sort_order=max_pos + 1,
            data={**snap, "id": ext_id},
        ))
        await _log_change(db, user, "create", "rule", tmpl.name, ext_id,
                          folder.device_group_id, f"From template: {tmpl.name}")
        await db.commit()
        return JSONResponse({"status": "ok", "ext_id": ext_id})
    except Exception as e:
        logger.error(f"Template apply failed: {e}")
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)
    finally:
        await client.close()


@router.post("/api/v1/templates/delete")
async def delete_templates(request: Request, data: TemplateDeleteRequest, db: AsyncSession = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error", "message": "Unauthorized"}, status_code=401)
    await db.execute(sa_delete(RuleTemplate).where(RuleTemplate.id.in_(data.template_ids)))
    await db.commit()
    return JSONResponse({"status": "ok", "deleted": len(data.template_ids)})


# =====================================================================
# BLOCK: Scheduler — auto-sync tasks
# =====================================================================

class SchedulerCreateRequest(BaseModel):
    task_type:       str = "sync"
    device_group_id: Optional[str] = None
    device_name:     Optional[str] = None
    interval_hours:  int = 24


class SchedulerToggleRequest(BaseModel):
    task_id: int
    enabled: bool


def _task_to_dict(t: ScheduledTask) -> dict:
    return {
        "id":              t.id,
        "task_type":       t.task_type,
        "device_group_id": t.device_group_id,
        "device_name":     t.device_name or t.device_group_id or "All devices",
        "interval_hours":  t.interval_hours,
        "enabled":         t.enabled,
        "status":          t.status,
        "last_error":      t.last_error,
        "last_run":        t.last_run.isoformat() if t.last_run else None,
        "next_run":        t.next_run.isoformat() if t.next_run else None,
        "created_at":      t.created_at.isoformat() if t.created_at else None,
        "created_by":      t.created_by,
    }


@router.get("/api/v1/scheduler/tasks")
async def scheduler_list(request: Request, db: AsyncSession = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error", "message": "Unauthorized"}, status_code=401)

    from app.config import SETTINGS
    tasks = (await db.execute(
        select(ScheduledTask).order_by(ScheduledTask.created_at.asc())
    )).scalars().all()

    return JSONResponse({
        "status": "ok",
        "tasks":  [_task_to_dict(t) for t in tasks],
        "has_credentials": bool(SETTINGS.NGFW_URL and SETTINGS.NGFW_USER and SETTINGS.NGFW_PASSWORD),
    })


@router.post("/api/v1/scheduler/tasks")
async def scheduler_create(request: Request, data: SchedulerCreateRequest, db: AsyncSession = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error", "message": "Unauthorized"}, status_code=401)

    if data.interval_hours < 1:
        return JSONResponse({"status": "error", "message": "interval_hours must be >= 1"}, status_code=400)

    next_run = datetime.now(timezone.utc) + timedelta(hours=data.interval_hours)
    task = ScheduledTask(
        task_type       = data.task_type,
        device_group_id = data.device_group_id or None,
        device_name     = data.device_name or None,
        interval_hours  = data.interval_hours,
        enabled         = True,
        status          = "idle",
        next_run        = next_run,
        created_by      = user.get("username"),
    )
    db.add(task)
    await db.commit()
    await db.refresh(task)
    return JSONResponse({"status": "ok", "task": _task_to_dict(task)})


@router.patch("/api/v1/scheduler/tasks/{task_id}")
async def scheduler_toggle(task_id: int, request: Request, data: SchedulerToggleRequest, db: AsyncSession = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error", "message": "Unauthorized"}, status_code=401)

    task = await db.get(ScheduledTask, task_id)
    if not task:
        return JSONResponse({"status": "error", "message": "Not found"}, status_code=404)

    task.enabled = data.enabled
    if data.enabled and task.next_run is None:
        task.next_run = datetime.now(timezone.utc) + timedelta(hours=task.interval_hours)
    await db.commit()
    return JSONResponse({"status": "ok", "task": _task_to_dict(task)})


@router.delete("/api/v1/scheduler/tasks/{task_id}")
async def scheduler_delete(task_id: int, request: Request, db: AsyncSession = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error", "message": "Unauthorized"}, status_code=401)

    task = await db.get(ScheduledTask, task_id)
    if not task:
        return JSONResponse({"status": "error", "message": "Not found"}, status_code=404)

    await db.delete(task)
    await db.commit()
    return JSONResponse({"status": "ok"})


@router.post("/api/v1/scheduler/tasks/{task_id}/run")
async def scheduler_run_now(task_id: int, request: Request, db: AsyncSession = Depends(get_db)):
    """Manually trigger a scheduled task immediately."""
    user = get_current_user(request)
    if not user:
        return JSONResponse({"status": "error", "message": "Unauthorized"}, status_code=401)

    task = await db.get(ScheduledTask, task_id)
    if not task:
        return JSONResponse({"status": "error", "message": "Not found"}, status_code=404)
    if task.status == "running":
        return JSONResponse({"status": "error", "message": "Task already running"}, status_code=409)

    from app.services.scheduler_service import run_scheduled_task
    asyncio.create_task(run_scheduled_task(task_id))
    return JSONResponse({"status": "ok", "message": "Task triggered"})
